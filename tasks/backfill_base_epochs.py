import uuid

import pandas as pd
import openpyxl

from cfm_api.app.utils import business_loader, excel_loader


def _uuid() -> str:
    return uuid.uuid4().hex


def _safe_float(val) -> float:
    try:
        if val in (None, ""):
            return 0.0
        return abs(float(val))
    except Exception:
        return 0.0


def _pick_epoch_seed(rows: pd.DataFrame) -> pd.Series:
    if rows.empty:
        return pd.Series(dtype=object)
    tagged = rows.copy()
    tagged["tag_norm"] = tagged.get("tag").astype(str).str.upper()
    opens = tagged[tagged["tag_norm"] == "OPEN"]
    if not opens.empty:
        return opens.sort_values("date").iloc[0]
    return tagged.sort_values("date").iloc[0]


def backfill_base_epochs() -> None:
    legs = business_loader.list_base_legs()
    if legs.empty:
        print("No base legs found; skipping base epoch backfill.")
        return

    epochs = business_loader.base_epochs_store.load()
    if epochs.empty:
        epochs = pd.DataFrame(columns=business_loader.base_epochs_store.fieldnames)

    position_map: dict[str, str] = {}
    if not epochs.empty and "position_id" in epochs.columns:
        epochs["position_id"] = epochs["position_id"].astype(str)
        position_map = (
            epochs.sort_values("start_date")
            .groupby("position_id")["base_epoch_id"]
            .last()
            .to_dict()
        )

    new_epochs = []
    for position_id, group in legs.groupby("position_id"):
        pid = str(position_id)
        if pid not in position_map:
            seed = _pick_epoch_seed(group)
            if seed.empty:
                continue
            epoch_id = _uuid()
            position_map[pid] = epoch_id
            cost_basis = _safe_float(seed.get("amount")) + _safe_float(seed.get("fees"))
            new_epochs.append(
                {
                    "base_epoch_id": epoch_id,
                    "position_id": pid,
                    "start_date": seed.get("date"),
                    "end_date": None,
                    "base_cost_basis_locked": cost_basis,
                    "entry_base_leg_id": seed.get("base_leg_id"),
                    "note": "backfill",
                }
            )

    if new_epochs:
        epochs = pd.concat([epochs, pd.DataFrame(new_epochs)], ignore_index=True)
        business_loader.base_epochs_store.overwrite(epochs.where(pd.notna(epochs), None).to_dict("records"))

    legs["position_id"] = legs["position_id"].astype(str)
    if "base_epoch_id" not in legs.columns:
        legs["base_epoch_id"] = None
    for idx, row in legs.iterrows():
        if row.get("base_epoch_id"):
            continue
        epoch_id = position_map.get(str(row.get("position_id") or ""))
        if epoch_id:
            legs.at[idx, "base_epoch_id"] = epoch_id
    business_loader.legs_store.overwrite(legs.where(pd.notna(legs), None).to_dict("records"))

    leg_epoch_map = (
        legs.dropna(subset=["base_leg_id", "base_epoch_id"])
        .astype({"base_leg_id": "string", "base_epoch_id": "string"})
        .set_index("base_leg_id")["base_epoch_id"]
        .to_dict()
    )

    for descriptor in excel_loader._discover_accounts():
        wb = openpyxl.load_workbook(descriptor.path)
        ws = wb["Ledger"]
        excel_loader._ensure_ledger_headers(ws)
        cols = excel_loader._get_ledger_columns(ws)
        if "Base Epoch Id" not in cols:
            wb.save(descriptor.path)
            continue
        for row_idx in range(2, ws.max_row + 1):
            current = ws.cell(row=row_idx, column=cols["Base Epoch Id"]).value
            if current not in (None, ""):
                continue
            base_leg_id = ws.cell(row=row_idx, column=cols.get("Base Leg Id", 0)).value if cols.get("Base Leg Id") else None
            base_pos_id = ws.cell(row=row_idx, column=cols.get("Base Position Id", 0)).value if cols.get("Base Position Id") else None
            epoch_id = None
            if base_leg_id:
                epoch_id = leg_epoch_map.get(str(base_leg_id))
            if not epoch_id and base_pos_id:
                epoch_id = position_map.get(str(base_pos_id))
            if epoch_id:
                ws.cell(row=row_idx, column=cols["Base Epoch Id"], value=epoch_id)
        wb.save(descriptor.path)


if __name__ == "__main__":
    backfill_base_epochs()
