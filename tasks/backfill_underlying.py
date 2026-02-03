#!/usr/bin/env python3
"""
Backfill underlying prices in Juice_Ledger_*.xlsx using alpha cache/AV lookup.

Usage:
  .venv/bin/python tasks/backfill_underlying.py --account Christie --start 2026-01-01 --end 2026-01-31
  .venv/bin/python tasks/backfill_underlying.py --account Travis --start 2026-01-01 --end 2026-01-31 --dry-run
"""
from __future__ import annotations

import argparse
import json
from datetime import datetime
from pathlib import Path
from typing import Optional

import openpyxl
import pandas as pd

from cfm_api.app.utils import excel_loader


def _unwrap_alpha_cache_payload(payload: object) -> dict:
    if isinstance(payload, dict):
        if payload.get("__alpha_cache_format__") == "raw":
            series = payload.get("time_series")
            if isinstance(series, dict):
                return series
        for key, value in payload.items():
            if isinstance(key, str) and key.lower().startswith("time series") and isinstance(value, dict):
                return value
    return payload if isinstance(payload, dict) else {}


def _load_cache(ticker: str, month_str: str) -> dict:
    path = excel_loader.ALPHA_CACHE_DIR / f"{ticker}_{month_str}.json"
    if not path.exists():
        return {}
    try:
        data = json.loads(path.read_text())
    except Exception:
        return {}
    data = _unwrap_alpha_cache_payload(data)
    return data if isinstance(data, dict) else {}


def _parse_cache(data: dict) -> list[tuple[str, pd.Timestamp]]:
    parsed: list[tuple[str, pd.Timestamp]] = []
    for ts in data.keys():
        try:
            dt = pd.to_datetime(ts, errors="coerce")
        except Exception:
            dt = pd.NaT
        if not pd.isna(dt):
            parsed.append((ts, dt))
    return parsed


def _resolve_from_cache(
    data: dict, parsed: list[tuple[str, pd.Timestamp]], trade_ts: pd.Timestamp
) -> tuple[Optional[float], Optional[str], dict]:
    if not parsed:
        return None, None, {}
    # Ledger timestamps are Central; Alpha cache timestamps are Eastern.
    trade_ts = trade_ts + pd.Timedelta(hours=1)
    trade_day = trade_ts.date()
    same_day = [item for item in parsed if item[1].date() == trade_day]
    if not same_day:
        return None, None, {}
    chosen_ts = min(same_day, key=lambda item: abs(item[1] - trade_ts))[0]
    entry = data.get(chosen_ts, {}) if isinstance(data, dict) else {}
    try:
        open_val = float(entry.get("1. open") or entry.get("open"))
    except Exception:
        open_val = None
    try:
        close_val = float(entry.get("4. close") or entry.get("close") or entry.get("price"))
    except Exception:
        close_val = None
    try:
        high_val = float(entry.get("2. high") or entry.get("high"))
    except Exception:
        high_val = None
    try:
        low_val = float(entry.get("3. low") or entry.get("low"))
    except Exception:
        low_val = None
    ohlc = [v for v in (open_val, close_val, high_val, low_val) if v is not None]
    if len(ohlc) >= 2:
        return round(sum(ohlc) / len(ohlc), 4), chosen_ts, entry
    if close_val is not None:
        return close_val, chosen_ts, entry
    if open_val is not None:
        return open_val, chosen_ts, entry
    return None, chosen_ts, entry


def _parse_date(value: object) -> Optional[pd.Timestamp]:
    if value in (None, ""):
        return None
    try:
        return pd.to_datetime(value, errors="coerce")
    except Exception:
        return None


def _as_float(val: object) -> Optional[float]:
    try:
        if val in (None, ""):
            return None
        return float(val)
    except Exception:
        return None


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--account", required=True, help="Account name/label (e.g., Christie, Travis)")
    parser.add_argument("--start", required=True, help="Start date (YYYY-MM-DD)")
    parser.add_argument("--end", required=True, help="End date (YYYY-MM-DD)")
    parser.add_argument("--dry-run", action="store_true", help="Print changes without writing")
    parser.add_argument("--force", action="store_true", help="Write underlying even if unchanged")
    parser.add_argument("--debug", action="store_true", help="Print per-row lookup details")
    parser.add_argument("--trace", action="store_true", help="Print cache file + resolved underlying per row")
    parser.add_argument("--limit", type=int, default=0, help="Stop after N rows (0 = no limit)")
    parser.add_argument("--tolerance", type=float, default=0.0001, help="Tolerance for treating values as equal")
    args = parser.parse_args()

    start = pd.to_datetime(args.start, errors="coerce")
    end = pd.to_datetime(args.end, errors="coerce")
    if pd.isna(start) or pd.isna(end):
        raise SystemExit("Invalid start/end date")

    descriptor = excel_loader._resolve_account(args.account)
    path = Path(descriptor.path)
    if not path.exists():
        raise SystemExit(f"Ledger not found: {path}")

    wb = openpyxl.load_workbook(path)
    if "Ledger" not in wb.sheetnames:
        raise SystemExit("Ledger sheet missing")
    ws = wb["Ledger"]
    excel_loader._ensure_ledger_headers(ws)
    cols = excel_loader._get_ledger_columns(ws)

    # Prefer header-based columns, but fall back to fixed positions:
    # Date = 2nd column, Symbol = 4th column, Underlying = last column.
    date_col = cols.get("Date") or 2
    symbol_col = cols.get("Symbol") or 4
    strike_col = cols.get("Strike") or 6
    expiry_col = cols.get("Expiry") or 7
    side_col = cols.get("Side")
    underlying_col = cols.get("Underlying") or ws.max_column

    if not all([date_col, symbol_col, expiry_col, underlying_col]):
        raise SystemExit("Ledger missing required columns for backfill")

    total = 0
    updated = 0
    cache: dict[tuple[str, str], tuple[dict, list[tuple[str, pd.Timestamp]]]] = {}
    for r in range(2, ws.max_row + 1):
        trade_dt = ws.cell(row=r, column=date_col).value
        trade_ts = _parse_date(trade_dt)
        if trade_ts is None:
            continue
        if trade_ts.normalize() < start.normalize() or trade_ts.normalize() > end.normalize():
            continue
        ticker = ws.cell(row=r, column=symbol_col).value
        if not ticker:
            continue
        ticker = str(ticker).strip().upper()
        expiry_val = ws.cell(row=r, column=expiry_col).value
        expiry_ts = _parse_date(expiry_val)
        side_val = ws.cell(row=r, column=side_col).value if side_col else None
        strike_val = ws.cell(row=r, column=strike_col).value if strike_col else None

        total += 1
        current_underlying = _as_float(ws.cell(row=r, column=underlying_col).value)
        month_str = trade_ts.strftime("%Y-%m")
        key = (ticker, month_str)
        data, parsed = cache.get(key, ({}, []))
        if not parsed:
            data = _load_cache(ticker, month_str)
            parsed = _parse_cache(data)
            if not parsed:
                try:
                    excel_loader._refresh_alpha_cache(ticker, expiry_ts)
                except Exception:
                    pass
                data = _load_cache(ticker, month_str)
                parsed = _parse_cache(data)
            cache[key] = (data, parsed)
        resolved, chosen_ts, entry = _resolve_from_cache(data, parsed, trade_ts)
        if args.debug:
            print(
                f"row {r}: {ticker} {trade_ts} chosen_ts={chosen_ts} current={current_underlying} resolved={resolved}"
            )
        if args.trace:
            cache_path = excel_loader.ALPHA_CACHE_DIR / f"{ticker}_{month_str}.json"
            try:
                o = entry.get("1. open") or entry.get("open")
                h = entry.get("2. high") or entry.get("high")
                l = entry.get("3. low") or entry.get("low")
                c = entry.get("4. close") or entry.get("close") or entry.get("price")
            except Exception:
                o = h = l = c = None
            if chosen_ts is None:
                sample = parsed[0][1].date().isoformat() if parsed else None
                last = parsed[-1][1].date().isoformat() if parsed else None
                print(
                    f"row {r}: cache={cache_path} parsed={len(parsed)} sample_day={sample} last_day={last} trade_day={trade_ts.date().isoformat()}"
                )
            print(
                f"row {r}: cache={cache_path} chosen_ts={chosen_ts} open={o} high={h} low={l} close={c} resolved={resolved} date={trade_ts}"
            )

        if resolved is None:
            continue

        if (not args.force) and current_underlying is not None and abs(current_underlying - float(resolved)) <= args.tolerance:
            continue

        updated += 1
        if args.dry_run:
            print(f"row {r}: {ticker} {trade_ts} underlying {current_underlying} -> {resolved}")
        else:
            ws.cell(row=r, column=underlying_col, value=float(resolved))
        if total % 200 == 0:
            print(f"Processed {total} rows... updated {updated}")
        if args.limit and total >= args.limit:
            break

    if not args.dry_run:
        wb.save(path)
    wb.close()

    print(f"Scanned {total} rows, updated {updated} rows")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
