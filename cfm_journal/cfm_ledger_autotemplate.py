
import argparse
import csv
import json
from collections import defaultdict
from pathlib import Path
from datetime import datetime, timedelta
from typing import Optional
import shutil
import os
import importlib.util

# Lazy-load openpyxl to avoid heavy imports at module import time (helps server subprocess)
_openpyxl_loaded = False
def _ensure_openpyxl():
    """Import openpyxl and commonly used utilities on first use."""
    global _openpyxl_loaded, openpyxl, get_column_letter, PatternFill
    if _openpyxl_loaded:
        return
    import importlib
    openpyxl = importlib.import_module("openpyxl")
    from openpyxl.utils import get_column_letter as _gcl
    from openpyxl.styles import PatternFill as _PatternFill
    get_column_letter = _gcl
    PatternFill = _PatternFill
    _openpyxl_loaded = True
import time
from typing import Iterable

# Add Alpha Vantage import and installation check
try:
    import requests
    import json
except ImportError:
    raise SystemExit("Please install requests with: pip install requests")

# Schwab support is optional and should not slow down ledger operations.
# Avoid importing schwab/authlib at module import time because it's heavy and can hang or be interrupted.
SCHWAB_AVAILABLE = importlib.util.find_spec("schwab") is not None
auth = None
client = None
easy_client = None

def _ensure_schwab():
    """Import schwab modules only when Schwab commands are used."""
    global auth, client, easy_client, SCHWAB_AVAILABLE
    if not SCHWAB_AVAILABLE:
        raise SystemExit("Schwab API not available. Install with: pip install schwab-py")
    if auth is not None and client is not None and easy_client is not None:
        return
    try:
        from schwab import auth as _auth, client as _client
        from schwab.auth import easy_client as _easy_client
        auth = _auth
        client = _client
        easy_client = _easy_client
    except Exception as exc:
        SCHWAB_AVAILABLE = False
        raise SystemExit(f"Schwab API import failed: {exc}") from exc

# Alpha Vantage API configuration
ALPHA_VANTAGE_API_KEY = "OXVB6OE4LG9JETYB"
ALPHA_VANTAGE_BASE_URL = "https://www.alphavantage.co/query"
ALPHA_CACHE_DIR = Path(__file__).resolve().parent / "alpha_cache"
_ALPHA_VANTAGE_CACHE = {}

# Schwab API configuration
SCHWAB_TOKENS_FILE = "schwab_tokens.json"
SCHWAB_CREDENTIALS_FILE = "schwab_credentials.json"
SCHWAB_API_KEY = "JHcZjmlg9bZvtGKAwgfzzTPXXICmPJFsN7VCiqgtw2jbzamG"
SCHWAB_SECRET = "eXb9DFG8GCJlATzzVX8jrhAHZBaNiA0Z73EWrKo8UwKdnTUu7ZVtlPwdzBTEvVE8"
SCHWAB_BASE_URL = "https://api.schwabapi.com/trader/v1"
SCHWAB_ACCOUNTS = {
    "12902569": "Christie",
    "38391626": "Travis"
}

TEMPLATE_NAME = "Juice_Ledger.xlsx"  # looked up in current working directory

def convert_central_to_eastern(trade_time: datetime, log: bool = True) -> datetime:
    """
    Convert Central Time to Eastern Time (add 1 hour).
    Alpha Vantage uses US/Eastern, Thinkorswim uses Central.
    """
    eastern_time = trade_time + timedelta(hours=1)
    if log:
        print(f"dY? Timezone Conversion:")
        print(f"   Input (Central): {trade_time.strftime('%Y-%m-%d %H:%M:%S')}")
        print(f"   Output (Eastern): {eastern_time.strftime('%Y-%m-%d %H:%M:%S')}")
    return eastern_time

def get_schwab_tokens():
    """
    Get Schwab API tokens using direct OAuth flow.
    """
    _ensure_schwab()
    
    try:
        # Check if we have existing tokens
        if os.path.exists(SCHWAB_TOKENS_FILE):
            try:
                with open(SCHWAB_TOKENS_FILE, 'r') as f:
                    tokens = json.load(f)
                
                # Check if tokens are still valid (not expired)
                if 'expires_at' in tokens:
                    expires_at = datetime.fromtimestamp(tokens['expires_at'])
                    if datetime.now() < expires_at:
                        print("[OK] Using existing valid Schwab tokens")
                        return tokens
                    else:
                        print("[WARNING] Existing tokens expired, refreshing...")
                else:
                    print("[WARNING] Existing tokens missing expiration info, refreshing...")
            except Exception as e:
                print(f"[WARNING] Error reading existing tokens: {e}")
                print("Will start fresh authentication...")
        
        # Start OAuth flow
        print("[KEY] Starting Schwab OAuth authentication...")
        print("This will open your browser for authentication.")
        
        import webbrowser
        import urllib.parse
        
        # Generate state parameter
        import secrets
        state = secrets.token_urlsafe(32)
        
        # Build authorization URL
        auth_url = (
            f"https://api.schwabapi.com/v1/oauth/authorize"
            f"?response_type=code"
            f"&client_id={SCHWAB_API_KEY}"
            f"&redirect_uri={urllib.parse.quote('https://127.0.0.1:8182')}"
            f"&state={state}"
        )
        
        print(f"\n[LINK] Please open this URL in your browser:")
        print(f"{auth_url}")
        print(f"\nAfter logging in and authorizing, copy the ENTIRE redirect URL and paste it below.")
        
        # Get redirect URL from user
        redirect_url = input("\nRedirect URL> ").strip()
        
        # Parse the authorization code from the redirect URL
        parsed_url = urllib.parse.urlparse(redirect_url)
        query_params = urllib.parse.parse_qs(parsed_url.query)
        
        if 'code' not in query_params:
            raise Exception("No authorization code found in redirect URL")
        
        auth_code = query_params['code'][0]
        
        # Exchange authorization code for tokens
        print("[REFRESH] Exchanging authorization code for tokens...")
        
        token_response = requests.post(
            'https://api.schwabapi.com/v1/oauth/token',
            data={
                'grant_type': 'authorization_code',
                'code': auth_code,
                'redirect_uri': 'https://127.0.0.1:8182',
                'client_id': SCHWAB_API_KEY,
                'client_secret': SCHWAB_SECRET
            },
            headers={'Content-Type': 'application/x-www-form-urlencoded'}
        )
        
        if token_response.status_code != 200:
            raise Exception(f"Token exchange failed: {token_response.status_code} - {token_response.text}")
        
        tokens = token_response.json()
        
        # Calculate expiration time
        if 'expires_in' in tokens:
            expires_at = datetime.now().timestamp() + tokens['expires_in']
            tokens['expires_at'] = expires_at
        
        # Save tokens to file
        with open(SCHWAB_TOKENS_FILE, 'w') as f:
            json.dump(tokens, f, indent=2)
        
        print("[OK] OAuth authentication completed and tokens saved")
        return tokens
        
    except Exception as e:
        print(f"[ERROR] Error during OAuth authentication: {e}")
        raise

def get_schwab_orders(account_name: str = None, days_back: int = 30):
    """
    Get historical orders/trades from Schwab API with execution times.
    
    Args:
        account_name: Optional account name to filter (Travis/Christie)
        days_back: Number of days to look back for orders
    
    Returns:
        List of order dictionaries with execution times
    """
    try:
        # Get access token
        access_token = get_schwab_access_token()
        
        # Calculate date range
        end_date = datetime.now()
        start_date = end_date - timedelta(days=days_back)
        
        # Format dates for API
        start_str = start_date.strftime('%Y-%m-%d')
        end_str = end_date.strftime('%Y-%m-%d')
        
        print(f"[SIGNAL] Fetching orders from {start_str} to {end_str}...")
        
        # Get account numbers first
        accounts_url = f"{SCHWAB_BASE_URL}/accounts"
        headers = {
            'Authorization': f'Bearer {access_token}',
            'Accept': 'application/json'
        }
        
        accounts_response = requests.get(accounts_url, headers=headers)
        accounts_response.raise_for_status()
        accounts_data = accounts_response.json()
        
        all_orders = []
        
        for account_data in accounts_data:
            securities_account = account_data.get('securitiesAccount', {})
            account_number = securities_account.get('accountNumber', '')
            account_name_mapped = SCHWAB_ACCOUNTS.get(account_number, f"Account_{account_number}")
            
            # Filter by account name if specified
            if account_name and account_name_mapped != account_name:
                continue
            
            print(f"[SEARCH] Fetching orders for account {account_number} ({account_name_mapped})")
            
            # Get orders for this account
            orders_url = f"{SCHWAB_BASE_URL}/accounts/{account_number}/orders"
            orders_params = {
                'fromEnteredTime': start_date.isoformat(),
                'toEnteredTime': end_date.isoformat(),
                'status': 'FILLED'  # Only get filled orders
            }
            
            orders_response = requests.get(orders_url, params=orders_params, headers=headers)
            orders_response.raise_for_status()
            orders_data = orders_response.json()
            
            for order in orders_data:
                # Extract order details
                order_id = order.get('orderId', '')
                order_status = order.get('status', '')
                entered_time = order.get('enteredTime', '')
                
                # Process order legs (individual trades)
                order_legs = order.get('orderLegCollection', [])
                
                for leg in order_legs:
                    instrument = leg.get('instrument', {})
                    asset_type = instrument.get('assetType', '')
                    
                    # Only process option orders
                    if asset_type != 'OPTION':
                        continue
                    
                    symbol = instrument.get('symbol', '')
                    instruction = leg.get('instruction', '')  # BUY, SELL, etc.
                    quantity = leg.get('quantity', 0)
                    
                    # Process fills (actual executions)
                    fills = leg.get('fills', [])
                    
                    for fill in fills:
                        fill_quantity = fill.get('quantity', 0)
                        fill_price = fill.get('price', 0.0)
                        fill_time = fill.get('time', entered_time)  # Execution time
                        
                        # Parse execution time
                        try:
                            if fill_time:
                                # Convert to datetime object
                                exec_dt = datetime.fromisoformat(fill_time.replace('Z', '+00:00'))
                                # Convert to local time (Central)
                                exec_dt_local = exec_dt.astimezone()
                            else:
                                exec_dt_local = datetime.now()
                        except Exception:
                            exec_dt_local = datetime.now()
                        
                        # Parse option symbol
                        option_data = parse_schwab_option_symbol(symbol)
                        
                        order_data = {
                            'account_number': account_number,
                            'account_name': account_name_mapped,
                            'order_id': order_id,
                            'order_status': order_status,
                            'symbol': symbol,
                            'underlying': option_data['underlying'],
                            'strike': option_data['strike'],
                            'expiration': option_data['expiration'],
                            'put_call': instrument.get('putCall', ''),
                            'instruction': instruction,
                            'quantity': fill_quantity,
                            'price': fill_price,
                            'execution_time': exec_dt_local,
                            'execution_time_str': exec_dt_local.strftime('%Y-%m-%d %H:%M:%S'),
                            'market_value': fill_quantity * fill_price * 100  # Options are per 100 shares
                        }
                        
                        all_orders.append(order_data)
        
        print(f"[OK] Found {len(all_orders)} option trade executions")
        return all_orders
        
    except Exception as e:
        print(f"[ERROR] Error fetching Schwab orders: {e}")
        return []

def get_schwab_positions(account_name: str = None):
    """
    Get current positions from Schwab API using direct API calls.
    
    Args:
        account_name: Optional account name to filter (Travis/Christie)
    
    Returns:
        List of position dictionaries
    """
    try:
        tokens = get_schwab_tokens()
        
        print("[SIGNAL] Fetching positions from Schwab API...")
        
        # Get accounts
        accounts_response = requests.get(
            f"{SCHWAB_BASE_URL}/accounts",
            headers={
                'Authorization': f"Bearer {tokens['access_token']}",
                'Accept': 'application/json'
            }
        )
        
        if accounts_response.status_code != 200:
            raise Exception(f"Failed to get accounts: {accounts_response.status_code} - {accounts_response.text}")
        
        accounts_data = accounts_response.json()
        print(f"[CHART] Found {len(accounts_data)} Schwab accounts")
        
        all_positions = []
        
        for account in accounts_data:
            account_number = account.get('securitiesAccount', {}).get('accountNumber', '')
            account_name_mapped = SCHWAB_ACCOUNTS.get(account_number, f"Account_{account_number}")
            
            if account_name and account_name_mapped != account_name:
                continue
            
            print(f"[SEARCH] Processing account {account_number} ({account_name_mapped})")
            
            try:
                # Get positions for this account
                positions_response = requests.get(
                    f"{SCHWAB_BASE_URL}/accounts/{account_number}",
                    params={'fields': 'positions'},
                    headers={
                        'Authorization': f"Bearer {tokens['access_token']}",
                        'Accept': 'application/json'
                    }
                )
                
                if positions_response.status_code != 200:
                    print(f"[WARNING] Error getting positions for account {account_number}: {positions_response.status_code}")
                    continue
                
                account_data = positions_response.json()
                positions = account_data.get('securitiesAccount', {}).get('positions', [])
                
                for position in positions:
                    instrument = position.get('instrument', {})
                    asset_type = instrument.get('assetType', '')
                    
                    # Only process option positions
                    if asset_type != 'OPTION':
                        continue
                    
                    # Get position quantities
                    long_quantity = position.get('longQuantity', 0)
                    short_quantity = position.get('shortQuantity', 0)
                    net_quantity = long_quantity - short_quantity
                    
                    if net_quantity == 0:
                        continue
                    
                    # Parse option symbol
                    symbol = instrument.get('symbol', '')
                    option_data = parse_schwab_option_symbol(symbol)
                    
                    # Determine position type
                    if net_quantity > 0:
                        position_type = "LONG"
                        quantity = net_quantity
                    else:
                        position_type = "SHORT"
                        quantity = abs(net_quantity)
                    
                    # Determine if it's a leap (expiration > 1 year from now)
                    is_leap = False
                    if option_data['expiration']:
                        try:
                            exp_date = datetime.strptime(option_data['expiration'], '%Y-%m-%d')
                            one_year_from_now = datetime.now() + timedelta(days=365)
                            is_leap = exp_date > one_year_from_now
                        except:
                            pass
                    
                    position_data = {
                        'account_number': account_number,
                        'account_name': account_name_mapped,
                        'symbol': symbol,
                        'underlying': option_data['underlying'],
                        'strike': option_data['strike'],
                        'expiration': option_data['expiration'],
                        'put_call': instrument.get('putCall', ''),
                        'position_type': position_type,
                        'quantity': quantity,
                        'is_leap': is_leap,
                        'market_value': position.get('marketValue', 0),
                        'average_price': position.get('averagePrice', 0),
                        'current_day_pnl': position.get('currentDayProfitLoss', 0)
                    }
                    
                    all_positions.append(position_data)
                    
                    leap_indicator = " (LEAP)" if is_leap else ""
                    print(f"   [CHART] {option_data['underlying']} {option_data['strike']} {option_data['expiration']} - {position_type} {quantity} contracts{leap_indicator}")
                
            except Exception as e:
                print(f"[WARNING] Error getting positions for account {account_number}: {e}")
                continue
        
        print(f"[OK] Found {len(all_positions)} option positions")
        return all_positions
        
    except Exception as e:
        print(f"[ERROR] Error fetching Schwab positions: {e}")
        return []

def parse_schwab_option_symbol(symbol: str):
    """
    Parse Schwab option symbol to extract underlying, strike, expiry, right.
    
    Args:
        symbol: Option symbol from Schwab (e.g., "AAPL  250117C00190000")
    
    Returns:
        Dictionary with parsed components
    """
    import re
    
    # Handle OCC format: AAPL  250117C00190000
    occ_pattern = r'^([A-Z]+)\s*(\d{6})([CP])(\d{8})$'
    occ_match = re.match(occ_pattern, symbol.replace(' ', ''))
    
    if occ_match:
        underlying, exp_date, right, strike = occ_match.groups()
        
        # Parse expiration date (YYMMDD format)
        year = 2000 + int(exp_date[:2])
        month = int(exp_date[2:4])
        day = int(exp_date[4:6])
        expiration = f"{year:04d}-{month:02d}-{day:02d}"
        
        # Parse strike price (divide by 1000 for OCC format)
        strike_price = float(strike) / 1000.0
        
        return {
            'underlying': underlying,
            'right': right,
            'strike': strike_price,
            'expiration': expiration,
            'raw_symbol': symbol
        }
    
    # If no pattern matches, return raw symbol
    print(f"[WARNING] Could not parse Schwab symbol: {symbol}")
    return {
        'underlying': '',
        'right': '',
        'strike': 0.0,
        'expiration': '',
        'raw_symbol': symbol
    }

def _alpha_cache_path(symbol_upper: str, month_str: str) -> Path:
    ALPHA_CACHE_DIR.mkdir(parents=True, exist_ok=True)
    return ALPHA_CACHE_DIR / f"{symbol_upper}_{month_str}.json"


def _load_alpha_cache(symbol_upper: str, month_str: str) -> Optional[dict]:
    path = _alpha_cache_path(symbol_upper, month_str)
    if not path.exists():
        return None
    try:
        with path.open("r", encoding="utf-8") as f:
            payload = json.load(f)
    except Exception:
        return None
    if isinstance(payload, dict) and payload.get("__alpha_cache_format__") == "raw":
        return payload.get("time_series") or None
    if isinstance(payload, dict) and "Time Series (1min)" in payload:
        return payload.get("Time Series (1min)") or None
    return payload


def _save_alpha_cache(symbol_upper: str, month_str: str, data: dict, raw: bool = False):
    path = _alpha_cache_path(symbol_upper, month_str)
    payload = data
    if raw:
        payload = {
            "__alpha_cache_format__": "raw",
            "raw": data,
            "time_series": data.get("Time Series (1min)") if isinstance(data, dict) else None,
        }
    with path.open("w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)


def _find_intraday_minute(time_series: dict, eastern_time: datetime):
    target_date = eastern_time.strftime("%Y-%m-%d")
    target_time_str = eastern_time.strftime("%Y-%m-%d %H:%M:00")
    daily_data = {
        time_str: minute_data
        for time_str, minute_data in time_series.items()
        if time_str.startswith(target_date)
    }
    daily_count = len(daily_data)
    if daily_count == 0:
        return None, 0, None, False
    if target_time_str in daily_data:
        return daily_data[target_time_str], daily_count, target_time_str, True

    closest_time = None
    closest_data = None
    min_diff = float("inf")
    for time_str, minute_data in daily_data.items():
        try:
            data_time = datetime.strptime(time_str, "%Y-%m-%d %H:%M:%S")
        except ValueError:
            continue
        time_diff = abs((data_time - eastern_time).total_seconds())
        if time_diff < min_diff and time_diff <= 300:
            min_diff = time_diff
            closest_time = time_str
            closest_data = minute_data

    if closest_time:
        return closest_data, daily_count, closest_time, False

    return None, daily_count, None, False


def get_cached_alpha_price(symbol: str, trade_time: datetime) -> Optional[float]:
    if not symbol or not isinstance(trade_time, datetime) or trade_time.year < 1900:
        return None
    symbol_upper = symbol.upper()
    eastern_time = convert_central_to_eastern(trade_time, log=False)
    month_str = f"{eastern_time.year}-{eastern_time.month:02d}"
    time_series = _load_alpha_cache(symbol_upper, month_str)
    if not time_series:
        return None
    minute_data, daily_count, matched_time, exact = _find_intraday_minute(time_series, eastern_time)
    if minute_data is None:
        return None
    try:
        high_price = float(minute_data['2. high'])
        low_price = float(minute_data['3. low'])
    except (KeyError, ValueError):
        return None
    return round((high_price + low_price) / 2, 2)

def fetch_alpha_vantage_time_series(symbol: str, eastern_time: datetime):
    '''Fetch 1-minute intraday data for the symbol/month and cache it.'''
    symbol_upper = symbol.upper()
    target_date = eastern_time.strftime("%Y-%m-%d")
    month_str = f"{eastern_time.year}-{eastern_time.month:02d}"
    cache_key = (symbol_upper, month_str)

    if cache_key in _ALPHA_VANTAGE_CACHE:
        print(f"Using cached Alpha Vantage data for {symbol_upper} for {month_str}")
        return _ALPHA_VANTAGE_CACHE[cache_key]

    persisted_data = _load_alpha_cache(symbol_upper, month_str)
    if persisted_data:
        print(f"Using saved Alpha Vantage cache for {symbol_upper} {month_str}")
        _ALPHA_VANTAGE_CACHE[cache_key] = persisted_data
        return persisted_data

    print(f"Fetching {symbol_upper} data for {target_date} from Alpha Vantage...")
    target_time_str_preview = eastern_time.strftime("%Y-%m-%d %H:%M:00")

    params = {
        'function': 'TIME_SERIES_INTRADAY',
        'symbol': symbol_upper,
        'interval': '1min',
        'month': month_str,
        'outputsize': 'full',
        'apikey': ALPHA_VANTAGE_API_KEY
    }

    print("Alpha Vantage API Request:")
    print(f"  URL: {ALPHA_VANTAGE_BASE_URL}")
    print(f"  Function: {params['function']}")
    print(f"  Symbol: {params['symbol']}")
    print(f"  Interval: {params['interval']}")
    print(f"  Month: {params['month']}")
    print(f"  API Key: {ALPHA_VANTAGE_API_KEY[:8]}...{ALPHA_VANTAGE_API_KEY[-4:]}")

    full_url = (
        f"{ALPHA_VANTAGE_BASE_URL}?function={params['function']}"
        f"&symbol={params['symbol']}&interval={params['interval']}"
        f"&month={params['month']}&outputsize={params['outputsize']}"
        f"&apikey={params['apikey']}"
    )

    print("\\nReview before sending:")
    print("=" * 80)
    print(f"Full URL: {full_url}")
    print(f"Target Time (Eastern): {target_time_str_preview}")
    print("=" * 80)
    print("Alpha Vantage has a limited number of API calls per day! Proceeding automatically if cache is missing.")

    try:
        response = requests.get(ALPHA_VANTAGE_BASE_URL, params=params, timeout=30)
        response.raise_for_status()

        print(f"Response Status: {response.status_code}")
        data = response.json()
        _save_alpha_cache(symbol_upper, month_str, data, raw=True)

        print("API Response:")
        print(json.dumps(data, indent=2))

        if 'Error Message' in data:
            print(f"Alpha Vantage API error: {data['Error Message']}")
            return None
        if 'Note' in data:
            print(f"Alpha Vantage API limit reached: {data['Note']}")
            return None
        if 'Information' in data:
            print(f"Alpha Vantage API info: {data['Information']}")
            return None

        time_series = data.get('Time Series (1min)')
        if not time_series:
            print(f"No time series data found for {symbol_upper}")
            return None
        _ALPHA_VANTAGE_CACHE[cache_key] = time_series
        daily_count = sum(1 for time_str in time_series if time_str.startswith(target_date))
        if daily_count:
            print(f"Found {daily_count} data points for {symbol_upper} on {target_date}")
        else:
            print(f"Warning: No intraday bars found for {symbol_upper} on {target_date}")
        return time_series

    except requests.exceptions.RequestException as e:
        print(f"Network error fetching {symbol_upper}: {str(e)}")
    except json.JSONDecodeError as e:
        print(f"Error decoding Alpha Vantage response: {str(e)}")
    except Exception as e:
        print(f"Unexpected error fetching {symbol_upper}: {str(e)}")

    return None


def get_alpha_vantage_price(symbol: str, trade_time: datetime):
    '''Fetches historical stock price from Alpha Vantage for the requested time.'''
    eastern_time = convert_central_to_eastern(trade_time)
    time_series = fetch_alpha_vantage_time_series(symbol, eastern_time)

    if not time_series:
        return None

    target_time_str = eastern_time.strftime("%Y-%m-%d %H:%M:00")
    print(f"Looking for data at (Eastern): {target_time_str}")
    minute_data, daily_count, matched_time, exact = _find_intraday_minute(time_series, eastern_time)
    target_date = eastern_time.strftime("%Y-%m-%d")
    print(f"Data points available for {symbol.upper()} on {target_date}: {daily_count}")
    if minute_data is None:
        print(f"No intraday data available for {symbol.upper()} on {target_date}")
        return None
    if not exact:
        print("No exact match, searching closest time within 5 minutes...")
        if matched_time:
            try:
                diff_seconds = abs((datetime.strptime(matched_time, '%Y-%m-%d %H:%M:%S') - eastern_time).total_seconds())
            except ValueError:
                diff_seconds = 0
            print(f"Found closest match at {matched_time} (diff: {diff_seconds:.0f} seconds)")
        else:
            print(f"No data found within 5 minutes of {target_time_str}")
            return None

    high_price = float(minute_data['2. high'])
    low_price = float(minute_data['3. low'])
    avg_price = (high_price + low_price) / 2

    print(f"Price data for {symbol} -> High: ${high_price:.2f}, Low: ${low_price:.2f}, Avg: ${avg_price:.2f}")
    if symbol.upper() == 'NVDA' and (high_price < 100 or high_price > 200):
        print(f"Warning: NVDA price ${high_price:.2f} seems unusual (expected ~150-200)")

    return round(avg_price, 2)


def get_minute_avg(symbol: str, trade_time: datetime):
    """
    Fetches stock price from Alpha Vantage API.
    Returns the latest available price or None if not found.
    """
    return get_alpha_vantage_price(symbol, trade_time)

def calculate_juice(premium: float, underlying: float, strike: float) -> float:
    """
    Calculate juice (extrinsic value) from premium, underlying price, and strike.
    Juice = Premium - max(0, Underlying - Strike)
    """
    intrinsic_value = max(0, underlying - strike)
    juice = premium - intrinsic_value
    return round(juice, 2)

DEFAULT_CSV_LOG = Path("cfm_trades_log.csv")
CSV_COLUMNS = [
    "Account",
    "Strategy",
    "Date",
    "Action",
    "Symbol",
    "Side",
    "Contracts",
    "Strike",
    "Expiry",
    "Premium/Buyback",
    "Underlying",
    "Juice/Contract",
    "Signed Juice ($)",
    "Signed Juice (per 100)",
    "Key",
    "Notes",
]

def sanitize_note(note: Optional[str]) -> str:
    if not note:
        return ""
    cleaned = str(note).replace(",", ";").replace("\n", " ").replace("\r", " ")
    return cleaned.strip()

def ensure_csv_header(csv_path: Path):
    csv_path = Path(csv_path)
    csv_path.parent.mkdir(parents=True, exist_ok=True)
    if not csv_path.exists():
        with csv_path.open("w", newline="", encoding="utf-8") as f:
            csv.writer(f).writerow(CSV_COLUMNS)
        return
    with csv_path.open(newline="", encoding="utf-8") as f:
        rows = list(csv.reader(f))
    if not rows:
        existing_header = []
        data_rows = []
    else:
        existing_header = rows[0]
        data_rows = rows[1:]
    if existing_header == CSV_COLUMNS:
        return
    with csv_path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(CSV_COLUMNS)
        for row in data_rows:
            row_dict = dict(zip(existing_header, row))
            writer.writerow([row_dict.get(col, "") for col in CSV_COLUMNS])

def append_csv_row(csv_path: Path, row: dict):
    ensure_csv_header(csv_path)
    with csv_path.open("a", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow([row.get(col, "") for col in CSV_COLUMNS])

def format_decimal(value: Optional[float]) -> str:
    if value is None:
        return ""
    return f"{value:.2f}"

def calculate_intrinsic_value(side: str, underlying: Optional[float], strike: Optional[float]) -> Optional[float]:
    if underlying is None or strike is None:
        return None
    if side == "Call":
        return max(0.0, underlying - strike)
    return max(0.0, strike - underlying)

def log_cash_event(
    strategy: str,
    account: str,
    action: str,
    ticker: str,
    premium: float,
    contracts: int = 1,
    side: str = "Call",
    strike: Optional[float] = None,
    expiry: Optional[str] = None,
    underlying: Optional[float] = None,
    date: Optional[datetime] = None,
    note: Optional[str] = "",
    csv_path: Optional[str] = None,
):
    action_clean = action.capitalize()
    if action_clean not in {"Open", "Close"}:
        raise ValueError("Action must be 'Open' or 'Close'")
    side_clean = side.capitalize() if side else "Call"
    if side_clean not in {"Call", "Put"}:
        raise ValueError("Side must be Call or Put")
    csv_dest = Path(csv_path) if csv_path else DEFAULT_CSV_LOG
    event_date = date if date else datetime.today()
    intrinsic = calculate_intrinsic_value(side_clean, underlying, strike)
    juice_per_contract = None
    if intrinsic is not None:
        juice_per_contract = round(float(premium) - intrinsic, 2)
    signed_juice_per_contract = juice_per_contract
    if juice_per_contract is not None and action_clean == "Close":
        signed_juice_per_contract = -juice_per_contract
    signed_total = None
    signed_per_100 = None
    if signed_juice_per_contract is not None:
        signed_total = round(signed_juice_per_contract * int(contracts), 2)
        signed_per_100 = round(signed_total * 100, 2)
    expiry_str = ""
    if expiry:
        try:
            expiry_str = parse_date(expiry).strftime("%Y-%m-%d")
        except Exception:
            expiry_str = expiry
    row = {
        "Account": account,
        "Strategy": strategy,
        "Date": event_date.date().isoformat(),
        "Action": action_clean,
        "Symbol": ticker.upper(),
        "Side": side_clean,
        "Contracts": int(contracts),
        "Strike": format_decimal(strike),
        "Expiry": expiry_str,
        "Premium/Buyback": format_decimal(round(float(premium), 2)),
        "Underlying": format_decimal(underlying),
        "Juice/Contract": format_decimal(juice_per_contract),
        "Signed Juice ($)": format_decimal(signed_total),
        "Signed Juice (per 100)": format_decimal(signed_per_100),
        "Key": composite_key(
            ticker.upper(),
            strike,
            expiry,
            side_clean,
            action_clean
        ),
        "Notes": sanitize_note(note),
    }
    append_csv_row(csv_dest, row)
    print(f"[OK] Logged {strategy} {action_clean} {side_clean} event for {row['Symbol']} into {csv_dest}")
HEADERS = [
    "Account","Date","Action","Symbol","Contracts","Strike","Expiry",
    "Premium/Buyback","Underlying","Key","Side","Condition"
]
COLS = {h:i+1 for i,h in enumerate(HEADERS)}

def write_formulas(ws, r: int):
    _ensure_openpyxl()
    def C(name): return get_column_letter(COLS[name])
    ws.cell(row=r, column=COLS["Key"]).value = (
        f'=UPPER({C("Symbol")}{r}) & "|" & '
        f'IF({C("Strike")}{r}="","",IF(MOD({C("Strike")}{r},1)=0,TEXT({C("Strike")}{r},"0"),TEXT({C("Strike")}{r},"0.################"))) & "|" & '
        f'TEXT({C("Expiry")}{r},"yyyy-mm-dd") & "|" & '
        f'IF({C("Side")}{r}="", "CALL", UPPER({C("Side")}{r})) & "|" & '
        f'IF({C("Action")}{r}="", "OPEN", UPPER({C("Action")}{r}))'
    )
    ws.cell(row=r, column=COLS["Juice/Contract"]).value = (
        f'=LET('
        f'  isClose, UPPER({C("Action")}{r})="CLOSE",'
        f'  side, UPPER({C("Side")}{r}),'
        f'  intrinsic, IF(side="PUT", MAX(0, {C("Strike")}{r} - {C("Underlying")}{r}), MAX(0, {C("Underlying")}{r} - {C("Strike")}{r})),'
        f'  extrinsic, {C("Premium/Buyback")}{r} - intrinsic,'
        f'  IF(AND({C("Strike")}{r}<>"", {C("Underlying")}{r}<>""),'
        f'     IF(isClose, IF(extrinsic<0, ABS(extrinsic), -extrinsic), extrinsic),'
        f'     IF(isClose, IF({C("Premium/Buyback")}{r}<0, ABS({C("Premium/Buyback")}{r}), -{C("Premium/Buyback")}{r}), {C("Premium/Buyback")}{r})'
        f'  )'
        f')'
    )
    ws.cell(row=r, column=COLS["Signed Juice ($)"]).value = (
        f'=N({C("Juice/Contract")}{r}) * N({C("Contracts")}{r})'
    )
    ws.cell(row=r, column=COLS["Signed Juice (per 100)"]).value = f'={C("Signed Juice ($)")}{r}*100'


def make_fresh_ledger(path: Path):
    _ensure_openpyxl()
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Ledger"
    for i,h in enumerate(HEADERS, start=1):
        ws.cell(row=1, column=i, value=h)
    wb.save(path)

def ensure_account_workbook(account: str, file_arg: Optional[str]) -> Path:
    # Decide destination path
    dest = Path(file_arg) if file_arg else Path(f"Juice_Ledger_{account}.xlsx")
    if dest.exists():
        # Upgrade headers in existing workbook to include any new columns (e.g., Condition)
        try:
            _ensure_openpyxl()
            wb = openpyxl.load_workbook(dest)
            if "Ledger" in wb.sheetnames:
                ws = wb["Ledger"]
                # Ensure header row matches HEADERS; append missing headers
                existing_headers = [ws.cell(row=1, column=i).value for i in range(1, ws.max_column + 1)]
                missing = [h for h in HEADERS if h not in existing_headers]
                if missing:
                    # Place missing headers at the end to avoid shifting data
                    start_col = ws.max_column + 1
                    for idx, h in enumerate(missing):
                        ws.cell(row=1, column=start_col + idx, value=h)
                    wb.save(dest)
        except Exception:
            # If upgrade fails, continue without blocking (append_row will still write headers)
            pass
        return dest
    # If template exists in CWD, clone it; else build fresh
    template = Path(TEMPLATE_NAME)
    if template.exists():
        shutil.copyfile(template, dest)
    else:
        make_fresh_ledger(dest)
    return dest

def parse_date(s: Optional[str]) -> datetime:
    if not s:
        return datetime.today()
    # Try full datetime formats first
    for fmt in ("%Y-%m-%d %H:%M", "%Y-%m-%dT%H:%M", "%Y-%m-%d %H:%M:%S", "%Y-%m-%dT%H:%M:%S"):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            pass
    # Fallback to date-only formats
    for fmt in ("%Y-%m-%d","%m/%d/%Y","%Y/%m/%d","%m/%d/%y"):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            pass
    raise SystemExit(f"Invalid date: {s}. Use YYYY-MM-DD, MM/DD/YYYY, or YYYY-MM-DD HH:MM.")

def parse_datetime(date_str: Optional[str], time_str: Optional[str]) -> datetime:
    """
    Combine date and time strings into a datetime object.
    If date is missing, use today's date.
    If time is missing, use current local time.
    """
    if date_str:
        # Parse the date part
        date_obj = parse_date(date_str)
        if time_str:
            # Parse time and combine with date
            try:
                for tfmt in ("%H:%M:%S", "%H:%M"):
                    try:
                        time_obj = datetime.strptime(time_str, tfmt).time()
                        return datetime.combine(date_obj.date(), time_obj)
                    except ValueError:
                        continue
                raise ValueError
            except ValueError:
                raise SystemExit(f"Invalid time: {time_str}. Use HH:MM or HH:MM:SS format.")
        else:
            # Use current time if no time provided
            return datetime.combine(date_obj.date(), datetime.now().time())
    else:
        # Use today's date
        if time_str:
            try:
                time_obj = datetime.strptime(time_str, "%H:%M").time()
                return datetime.combine(datetime.today().date(), time_obj)
            except ValueError:
                raise SystemExit(f"Invalid time: {time_str}. Use HH:MM format.")
        else:
            # Use current datetime
            return datetime.now()

def prompt_value(prompt_text: str, default: Optional[str] = None, required: bool = False) -> str:
    """Prompt the user for a value with an optional default."""
    suffix = f" [{default}]" if default is not None else ""
    while True:
        raw = input(f"{prompt_text}{suffix}: ").strip()
        if raw:
            return raw
        if default is not None:
            return default
        if not required:
            return ""
        print("A value is required.")

def prompt_number(prompt_text: str, cast, default: Optional[float] = None, required: bool = True):
    """Prompt for a numeric value and cast it to int/float."""
    suffix = f" [{default}]" if default is not None else ""
    while True:
        raw = input(f"{prompt_text}{suffix}: ").strip()
        if not raw:
            if default is not None:
                return default
            if not required:
                return default
            print("A value is required.")
            continue
        try:
            return cast(raw)
        except (TypeError, ValueError):
            print("Please enter a valid number.")

def prompt_choice(prompt_text: str, choices, default: Optional[str] = None) -> str:
    """Prompt for a value that must be one of the provided choices."""
    choice_str = "/".join(choices)
    suffix = f" [{default}]" if default else ""
    while True:
        raw = input(f"{prompt_text} ({choice_str}){suffix}: ").strip()
        if not raw and default:
            return default
        for option in choices:
            if raw.lower() == option.lower():
                return option
        print(f"Please choose one of: {', '.join(choices)}")

def next_row(ws) -> int:
    return ws.max_row + 1 if ws.max_row >= 1 else 2

def is_file_open(file_path: Path) -> bool:
    """Check if a file is currently open by another process."""
    try:
        # Try to open the file in append mode
        with open(file_path, 'a'):
            return False
    except (IOError, OSError):
        return True

def handle_file_permission_error(file_path: Path, max_retries: int = 3) -> bool:
    """Handle file permission errors by asking user to close the file."""
    for attempt in range(max_retries):
        try:
            # Try to save the file
            return True
        except PermissionError:
            if attempt < max_retries - 1:
                print(f"\n[ERROR] Error: Cannot save to {file_path.name}")
                print("The file appears to be open in another program (like Excel).")
                print("\nPlease:")
                print("1. Close the Excel file if it's open")
                print("2. Make sure no other programs are using the file")
                print("3. Press Enter when ready to try again...")
                
                input("Press Enter to retry...")
            else:
                print(f"\n[ERROR] Failed to save after {max_retries} attempts.")
                print("Please close the Excel file and try again.")
                return False
    return False

def composite_key(
    symbol: str,
    strike: Optional[float],
    expiry: Optional[str],
    side: Optional[str],
    action: Optional[str] = None,
) -> str:
    exp = ""
    if expiry:
        try:
            if hasattr(expiry, "strftime"):
                exp = expiry.strftime("%Y-%m-%d")
            else:
                exp = parse_date(expiry).strftime("%Y-%m-%d")
        except SystemExit:
            exp = str(expiry)
        except Exception:
            exp = expiry
    s_strike = ""
    if strike is not None:
        # Format as integer if it's a whole number, otherwise keep decimals
        try:
            strike_float = float(strike)
            if strike_float.is_integer():
                s_strike = str(int(strike_float))
            else:
                s_strike = f"{strike_float:.6f}".rstrip("0").rstrip(".")
        except (ValueError, TypeError):
            s_strike = str(strike)
    side_part = (side or "Call").upper()
    action_part = (action or "Open").upper()
    return f"{symbol.upper()}|{s_strike}|{exp}|{side_part}|{action_part}"


def base_from_composite(key: str) -> str:
    return "|".join(key.split("|")[:4])


def batch_base_key(cleaned_row: dict) -> str:
    """Return the base key (symbol|strike|expiry|side) for batch stats."""
    def format_strike(value):
        if value is None:
            return ""
        try:
            strike_val = float(value)
            if strike_val.is_integer():
                return str(int(strike_val))
            return f"{strike_val:.6f}".rstrip("0").rstrip(".")
        except (ValueError, TypeError):
            return str(value)

    def normalize_expiry(value):
        if not value:
            return ""
        if hasattr(value, "strftime"):
            return value.strftime("%Y-%m-%d")
        try:
            return parse_date(str(value)).strftime("%Y-%m-%d")
        except Exception:
            return str(value).split(" ")[0]

    symbol = (cleaned_row.get("Symbol") or "").upper()
    strike = format_strike(cleaned_row.get("Strike"))
    expiry = normalize_expiry(cleaned_row.get("Expiry"))
    side = (cleaned_row.get("Side") or "Call").upper()
    return f"{symbol}|{strike}|{expiry}|{side}"


def collect_ledger_key_tuples(file_paths: list[Path]) -> list[tuple[str,str,str,str]]:
    tuples: list[tuple[str,str,str,str]] = []
    for ledger_path in file_paths:
        if not ledger_path or not ledger_path.exists():
            continue
        _ensure_openpyxl()
        wb = openpyxl.load_workbook(ledger_path, data_only=True)
        ws = wb["Ledger"]
        for r in range(2, ws.max_row + 1):
            symbol = (ws.cell(row=r, column=COLS["Symbol"]).value or "").strip()
            strike = ws.cell(row=r, column=COLS["Strike"]).value
            expiry = ws.cell(row=r, column=COLS["Expiry"]).value
            side = (ws.cell(row=r, column=COLS["Side"]).value or "Call").strip().upper()
            action_text = (ws.cell(row=r, column=COLS["Action"]).value or "Open").strip()
            action_upper = "CLOSE" if "close" in action_text.lower() else "OPEN"
            if not symbol or strike is None or expiry is None:
                continue
            tuples.append((
                composite_key(symbol, strike, expiry, side, action_upper),
                action_upper,
                (ws.cell(row=r, column=COLS["Account"]).value or "").strip().upper(),
                side
            ))
    return tuples


def log_ledger_positions(file_paths: list[Path]):
    for ledger_path in file_paths:
        if not ledger_path or not ledger_path.exists():
            continue
        positions = get_open_positions(ledger_path)
        print(f"Ledger open positions in {ledger_path.name}:")
        if not positions:
            print("  (none)")
            continue
        for key, qty in sorted(positions.items()):
            print(f"  {key}: {qty} contracts")


def summarize_batch_rows(rows: list[tuple]) -> dict:
    from collections import defaultdict

    actions_by_base = defaultdict(set)
    for _, _, _, _, _, _, action_label, base_key in rows:
        act = action_label.upper()
        actions_by_base[base_key].add(act)

    opens = {key for key, acts in actions_by_base.items() if "OPEN" in acts}
    closes = {key for key, acts in actions_by_base.items() if "CLOSE" in acts}
    paired = {key for key, acts in actions_by_base.items() if {"OPEN", "CLOSE"} <= acts}
    only_opens = opens - paired
    only_closes = closes - paired
    return {
        "opens": opens,
        "closes": closes,
        "paired": paired,
        "only_opens": only_opens,
        "only_closes": only_closes,
    }


def format_keys(keys: Iterable[str]) -> str:
    return ", ".join(sorted(keys)) if keys else "(none)"


def log_batch_stats(summary: dict, csv_path: Path):
    print(f"Batch stats for {csv_path.name}:")
    print(f"  Keys opened in batch: {len(summary['opens'])}")
    for key in sorted(summary['opens']):
        print(f"    Open: {key}")
    print(f"  Keys closed in batch: {len(summary['closes'])}")
    for key in sorted(summary['closes']):
        print(f"    Close: {key}")
    print(f"  Paired open/close keys in batch: {len(summary['paired'])}")
    for key in sorted(summary['paired']):
        print(f"    Paired: {key}")


def log_batch_keys(rows: list[tuple]) -> None:
    print("Batch-derived keys:")
    for _, _, row_num, command, command_raw, _, action_label, base_key in rows:
        key = f"{base_key}|{action_label.upper()}"
        print(f"  Row {row_num}: {command_raw} -> {key}")

def current_open_contracts(wb, account: str, key: str, file_path: Path) -> int:
    # Load the workbook normally to get the raw data, then manually calculate the key
    _ensure_openpyxl()
    _ensure_openpyxl()
    wb_data = openpyxl.load_workbook(file_path, data_only=False)
    ws = wb_data["Ledger"]
    total = 0
    for r in range(2, ws.max_row+1):
        acc = (ws.cell(row=r, column=COLS["Account"]).value or "").strip()
        symbol = (ws.cell(row=r, column=COLS["Symbol"]).value or "").strip().upper()
        strike = ws.cell(row=r, column=COLS["Strike"]).value
        expiry = ws.cell(row=r, column=COLS["Expiry"]).value
        
        raw_side = ws.cell(row=r, column=COLS["Side"]).value
        if raw_side in (None, ""):
            side = "Call"
        else:
            side = str(raw_side).strip() or "Call"
        # Manually calculate the key from the raw data
        action_text = (ws.cell(row=r, column=COLS["Action"]).value or "").strip()
        action_value = "CLOSE" if "close" in action_text.lower() else "OPEN"
        if strike is not None and expiry is not None:
            # Format strike price properly
            if isinstance(strike, (int, float)):
                if strike == int(strike):
                    strike_str = str(int(strike))
                else:
                    strike_str = f"{strike:.6f}".rstrip("0").rstrip(".")
            else:
                strike_str = str(strike)
            
            # Format expiry date
            if hasattr(expiry, 'strftime'):
                expiry_str = expiry.strftime("%Y-%m-%d")
            else:
                try:
                    expiry_str = parse_date(str(expiry)).strftime("%Y-%m-%d")
                except (SystemExit, Exception):
                    expiry_str = str(expiry).split(" ")[0]
            
            keyv = composite_key(symbol, strike, expiry_str, side, action_value)
        else:
            keyv = ""
        
        action = action_value
        contracts = ws.cell(row=r, column=COLS["Contracts"]).value
        if acc == account and keyv == key and contracts is not None:
            if action == "OPEN":
                total += int(contracts)
            elif action == "CLOSE":
                total -= int(contracts)
    return total

def get_open_positions(file_path: Path, account: str = None) -> dict:
    """Get all open positions for an account or all accounts.
    Returns a dictionary with position keys and their net contract counts."""
    _ensure_openpyxl()
    wb_data = openpyxl.load_workbook(file_path, data_only=False)
    ws = wb_data["Ledger"]
    
    position_totals = {}
    
    for r in range(2, ws.max_row + 1):
        acc = (ws.cell(row=r, column=COLS["Account"]).value or "").strip()
        symbol = (ws.cell(row=r, column=COLS["Symbol"]).value or "").strip().upper()
        strike = ws.cell(row=r, column=COLS["Strike"]).value
        expiry = ws.cell(row=r, column=COLS["Expiry"]).value
        action = (ws.cell(row=r, column=COLS["Action"]).value or "").strip().upper()
        contracts = ws.cell(row=r, column=COLS["Contracts"]).value
        
        # Skip if account filter doesn't match
        if account and acc != account:
            continue
            
        # Skip if no contracts or invalid data
        if contracts is None or not symbol or strike is None or expiry is None:
            continue
            
        # Create position key
        if isinstance(strike, (int, float)):
            if strike == int(strike):
                strike_str = str(int(strike))
            else:
                strike_str = f"{strike:.6f}".rstrip("0").rstrip(".")
        else:
            strike_str = str(strike)
            
        if hasattr(expiry, 'strftime'):
            expiry_str = expiry.strftime("%Y-%m-%d")
        else:
            expiry_str = str(expiry)
            
        key = f"{acc}|{symbol}|{strike_str}|{expiry_str}"
        
        # Track net contracts
        if key not in position_totals:
            position_totals[key] = 0
            
        if action == "OPEN":
            position_totals[key] += int(contracts)
        elif action == "CLOSE":
            position_totals[key] -= int(contracts)
    
    # Return only positions with net contracts > 0
    return {k: v for k, v in position_totals.items() if v > 0}

def highlight_open_positions(file_path: Path, account: str = None):
    """Highlight rows that represent open positions in the Excel file."""
    _ensure_openpyxl()
    wb = openpyxl.load_workbook(file_path)
    ws = wb["Ledger"]
    
    # Get open positions
    open_positions = get_open_positions(file_path, account)
    
    # Create highlight fill
    _ensure_openpyxl()
    highlight_fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
    
    # Clear existing highlighting first
    for r in range(2, ws.max_row + 1):
        for c in range(1, len(HEADERS) + 1):
            cell = ws.cell(row=r, column=c)
            try:
                if cell.fill.start_color.rgb == "FFFF99":
                    cell.fill = PatternFill()
            except Exception:
                # Some cells may not have start_color or rgb; ignore
                pass
    
    # Highlight rows with open positions
    highlighted_count = 0
    for r in range(2, ws.max_row + 1):
        acc = (ws.cell(row=r, column=COLS["Account"]).value or "").strip()
        symbol = (ws.cell(row=r, column=COLS["Symbol"]).value or "").strip().upper()
        strike = ws.cell(row=r, column=COLS["Strike"]).value
        expiry = ws.cell(row=r, column=COLS["Expiry"]).value
        action = (ws.cell(row=r, column=COLS["Action"]).value or "").strip().upper()
        
        # Skip if account filter doesn't match
        if account and acc != account:
            continue
            
        # Skip if no valid data
        if not symbol or strike is None or expiry is None or action not in ["OPEN", "CLOSE"]:
            continue
            
        # Create position key
        if isinstance(strike, (int, float)):
            if strike == int(strike):
                strike_str = str(int(strike))
            else:
                strike_str = f"{strike:.6f}".rstrip("0").rstrip(".")
        else:
            strike_str = str(strike)
            
        if hasattr(expiry, 'strftime'):
            expiry_str = expiry.strftime("%Y-%m-%d")
        else:
            expiry_str = str(expiry)
            
        key = f"{acc}|{symbol}|{strike_str}|{expiry_str}"
        
        # Check if this position is currently open
        if key in open_positions:
            # Highlight the entire row
            for c in range(1, len(HEADERS) + 1):
                ws.cell(row=r, column=c).fill = highlight_fill
            highlighted_count += 1
    
    # Save the file
    try:
        wb.save(file_path)
        print(f"[OK] Highlighted {highlighted_count} rows with open positions in {file_path.name}")
        if open_positions:
            print("Open positions found:")
            for key, contracts in open_positions.items():
                acc, symbol, strike, expiry = key.split("|", 3)
                print(f"  {acc}: {contracts} contracts of {symbol} {strike} {expiry}")
        else:
            print("No open positions found.")
    except PermissionError:
        print(f"[ERROR] Error: Cannot save to {file_path.name}")
        print("The file appears to be open in another program (like Excel).")
        print("Please close the file and try again.")
        return False
    
    return True

def write_formulas(ws, r: int):
    # No-op: formulas removed to avoid adding unused columns
    return

def append_row(path: Path, row: dict):
    _ensure_openpyxl()
    # Ensure per-account workbook exists (clone template or create fresh)
    # (If path already exists, it's returned unchanged)
    path = ensure_account_workbook(row["Account"], str(path))
    
    # Check if file is open before proceeding
    if is_file_open(path):
        print(f"\n[WARNING] Warning: {path.name} appears to be open in another program.")
        print("This may cause a permission error when trying to save.")
        print("Please close the file in Excel or other programs before continuing.")
        input("Press Enter when ready to continue...")
    
    wb = openpyxl.load_workbook(path)
    ws = wb["Ledger"]

    # Ensure required headers (Key, Side, Condition) exist; avoid adding optional ones that could misalign data
    header_row = [ws.cell(row=1, column=i).value for i in range(1, ws.max_column + 1)]
    required_headers = ["Key", "Side", "Condition"]
    for req in required_headers:
        if req not in header_row:
            ws.cell(row=1, column=ws.max_column + 1, value=req)
            header_row = [ws.cell(row=1, column=i).value for i in range(1, ws.max_column + 1)]
    header_to_col = {h: idx + 1 for idx, h in enumerate(header_row) if h}

    # Compute Key if missing
    if not row.get("Key"):
        row["Key"] = composite_key(
            row.get("Symbol"),
            row.get("Strike"),
            row.get("Expiry"),
            row.get("Side"),
            row.get("Action"),
        )

    r = next_row(ws)
    for k, v in row.items():
        if k in header_to_col:
            ws.cell(row=r, column=header_to_col[k], value=v)
    # Formulas for juice are not used; skip to avoid missing-column errors
    
    # Try to save with error handling
    max_retries = 3
    for attempt in range(max_retries):
        try:
            wb.save(path)
            print(f"[OK] Added row {r} to {path.name}.")
            break
        except PermissionError:
            if attempt < max_retries - 1:
                print(f"\n[ERROR] Error: Cannot save to {path.name}")
                print("The file appears to be open in another program (like Excel).")
                print("\nPlease:")
                print("1. Close the Excel file if it's open")
                print("2. Make sure no other programs are using the file")
                print("3. Press Enter when ready to try again...")
                
                input("Press Enter to retry...")
            else:
                print(f"\n[ERROR] Failed to save after {max_retries} attempts.")
                print("Please close the Excel file and try again.")
                raise SystemExit("Cannot save file - it appears to be open in another program.")

def cmd_open(args):
    dest = ensure_account_workbook(args.account, args.file)
    # Ensure strike price is stored with proper precision
    strike_value = round(float(args.strike), 6) if args.strike is not None else None
    side_clean = (args.side or "Call").capitalize()
    
    # Parse datetime from date and time arguments
    trade_time = parse_datetime(args.date, args.time)
    print(f"[DATE] Trade time being used: {trade_time}")
    print(f"[DATE] Date: {trade_time.date()}")
    print(f"[DATE] Time: {trade_time.time()}")
    
    # Auto-fetch underlying price if requested and not provided
    underlying_price = args.underlying
    if args.auto_price and args.underlying is None:
        print(f"[REFRESH] Attempting to fetch {args.symbol} historical price from Alpha Vantage...")
        print(f"[DATE] Fetching 1-minute data for {trade_time.strftime('%Y-%m-%d %H:%M')}")
        avg_price = get_minute_avg(args.symbol, trade_time)
        if avg_price:
            underlying_price = avg_price
        else:
            print(f"[ERROR] Auto-fetch failed for {args.symbol}")
            print(f"[INFO] Alpha Vantage API appears to be having issues or rate limit reached.")
            print(f"[INFO] Please run the command again with --underlying <price> to manually specify the stock price.")
            print(f"[INFO] Example: --underlying 153.42")
            raise SystemExit("Auto-fetch failed. Please provide --underlying price manually.")
    
    # Validate that we have underlying price for juice calculation
    if underlying_price is None:
        raise SystemExit("Underlying price is required for juice calculation. Provide --underlying or use --auto-price.")
    
    # Calculate juice from premium, underlying, and strike
    calculated_juice = calculate_juice(args.premium, underlying_price, args.strike)
    print(f"Calculated juice: {calculated_juice} (Premium: {args.premium}, Underlying: {underlying_price}, Strike: {args.strike})")
    
    row = {
        "Account": args.account,
        "Date": trade_time,
        "Action": "OPEN",
        "Symbol": args.symbol.upper(),
        "Side": side_clean,
        "Contracts": int(args.contracts),
        "Strike": strike_value,
        "Expiry": parse_date(args.expiry) if args.expiry else None,
        "Premium/Buyback": args.premium,
        "Underlying": underlying_price,
        "Condition": args.condition or "",
        "Notes": args.notes or ""
    }
    row["Key"] = composite_key(
        row["Symbol"],
        row["Strike"],
        args.expiry,
        row["Side"],
        row["Action"]
    )
    append_row(dest, row)
    log_cash_event(
        strategy="CFM_CC",
        account=args.account,
        action="Open",
        ticker=args.symbol,
        premium=args.premium,
        side=side_clean,
        strike=strike_value,
        expiry=args.expiry,
        underlying=underlying_price,
        contracts=int(args.contracts),
        date=trade_time,
        note=args.notes or "",
        csv_path=args.csv_log,
    )

def cmd_close(args):
    dest = ensure_account_workbook(args.account, args.file)
    _ensure_openpyxl()
    wb = openpyxl.load_workbook(dest)
    # Ensure strike price is stored with proper precision
    strike_value = round(float(args.strike), 6) if args.strike is not None else None
    side_clean = (args.side or "Call").capitalize()
    key = composite_key(args.symbol, strike_value, args.expiry, side_clean, "Close")
    open_key = composite_key(args.symbol, strike_value, args.expiry, side_clean, "Open")
    avail = current_open_contracts(wb, args.account, open_key, dest)
    qty = int(args.contracts)
    if avail < qty:
        raise SystemExit(f"Trying to close {qty} but only {avail} contracts are open for {args.account} {key}.")
    
    # Parse datetime from date and time arguments
    trade_time = parse_datetime(args.date, args.time)
    
    # Auto-fetch underlying price if requested and not provided
    underlying_price = args.underlying_close
    if args.auto_price and args.underlying_close is None:
        print(f"[REFRESH] Attempting to fetch {args.symbol} historical price from Alpha Vantage...")
        print(f"[DATE] Fetching 1-minute data for {trade_time.strftime('%Y-%m-%d %H:%M')}")
        avg_price = get_minute_avg(args.symbol, trade_time)
        if avg_price:
            underlying_price = avg_price
        else:
            print(f"[ERROR] Auto-fetch failed for {args.symbol}")
            print(f"[INFO] Alpha Vantage API appears to be having issues or rate limit reached.")
            print(f"[INFO] Please run the command again with --underlying-close <price> to manually specify the stock price.")
            print(f"[INFO] Example: --underlying-close 153.42")
            raise SystemExit("Auto-fetch failed. Please provide --underlying-close price manually.")
    
    # Validate that we have underlying price for juice calculation
    if underlying_price is None:
        raise SystemExit("Underlying price is required for juice calculation. Provide --underlying-close or use --auto-price.")
    
    # Calculate juice from buyback premium, underlying, and strike
    calculated_juice = calculate_juice(args.buyback, underlying_price, args.strike)
    print(f"Calculated juice: {calculated_juice} (Buyback: {args.buyback}, Underlying: {underlying_price}, Strike: {args.strike})")
    
    row = {
        "Account": args.account,
        "Date": trade_time,
        "Action": "CLOSE",
        "Symbol": args.symbol.upper(),
        "Side": side_clean,
        "Contracts": qty,
        "Strike": strike_value,
        "Expiry": parse_date(args.expiry) if args.expiry else None,
        "Premium/Buyback": args.buyback,
        "Underlying": underlying_price,
        "Condition": args.condition or "",
        "Notes": args.notes or ""
    }
    row["Key"] = key
    append_row(dest, row)
    log_cash_event(
        strategy="CFM_CC",
        account=args.account,
        action="Close",
        ticker=args.symbol,
        premium=args.buyback,
        side=side_clean,
        strike=strike_value,
        expiry=args.expiry,
        underlying=underlying_price,
        contracts=qty,
        date=trade_time,
        note=args.notes or "",
        csv_path=args.csv_log,
    )

def cmd_jl(args):
    event_date = parse_datetime(args.date, args.time)
    underlying_price = args.underlying
    if args.auto_price and underlying_price is None:
        print(f"🔄 Attempting to fetch {args.ticker} price for {event_date.strftime('%Y-%m-%d %H:%M')}...")
        fetched = get_minute_avg(args.ticker, event_date)
        if fetched:
            underlying_price = fetched
            print(f"✅ Auto-fetched price: {underlying_price}")
        else:
            raise SystemExit("Auto-fetch failed. Provide --underlying or try again later.")

    log_cash_event(
        strategy="JL",
        account=args.account,
        action=args.action,
        ticker=args.ticker,
        premium=args.premium,
        side=args.side,
        strike=args.strike,
        expiry=args.expiry,
        underlying=underlying_price,
        contracts=args.contracts,
        date=event_date,
        note=args.note or "",
        csv_path=args.csv_log,
    )

def cmd_jl_summary(args):
    csv_dest = Path(args.csv_log) if args.csv_log else DEFAULT_CSV_LOG
    if not csv_dest.exists():
        print(f"⚠️  CSV log not found at {csv_dest}")
        return

    totals = defaultdict(float)
    with csv_dest.open(newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            if row.get("Strategy") != "JL":
                continue
            date_str = row.get("Date")
            if not date_str:
                continue
            try:
                dt = datetime.strptime(date_str, "%Y-%m-%d")
            except ValueError:
                continue
            week_key = dt.isocalendar()[1]
            try:
                amount = float(row.get("Signed Juice ($)", "0"))
            except ValueError:
                continue
            totals[week_key] += amount

    if not totals:
        print(f"📭 No JL rows found in {csv_dest}")
        return

    print(f"\n📊 JL net juice totals by week ({csv_dest}):")
    print("Week NetJuice")
    print("---- --------")

    for week in sorted(totals.keys()):
        print(f"{week:>4} {totals[week]:>8.2f}")

def cmd_highlight(args):
    """Highlight open positions in the Excel file."""
    if args.file:
        file_path = Path(args.file)
    else:
        if args.account:
            file_path = Path(f"Juice_Ledger_{args.account}.xlsx")
        else:
            # If no account specified, try to find any ledger file
            possible_files = [Path("Juice_Ledger_Travis.xlsx"), Path("Juice_Ledger_Christie.xlsx"), Path("Juice_Ledger.xlsx")]
            file_path = None
            for pf in possible_files:
                if pf.exists():
                    file_path = pf
                    break
            if not file_path:
                raise SystemExit("No ledger file found. Please specify --file or --account.")
    
    if not file_path.exists():
        raise SystemExit(f"File {file_path} does not exist.")
    
    highlight_open_positions(file_path, args.account)

def cmd_sync(args):
    """Sync positions from Schwab to Excel ledger."""
    _ensure_schwab()
    
    print("🔄 Syncing positions from Schwab to Excel ledger...")
    
    # Get positions from Schwab
    schwab_positions = get_schwab_positions(args.account)
    
    if not schwab_positions:
        print("📭 No option positions found in Schwab accounts")
        return
    
    print(f"📊 Found {len(schwab_positions)} option positions in Schwab")
    
    # Process each position
    synced_count = 0
    leap_count = 0
    short_call_count = 0
    
    for position in schwab_positions:
        try:
            # Skip if we can't parse the symbol
            if not position['underlying']:
                print(f"⚠️  Skipping unparseable symbol: {position['symbol']}")
                continue
            
            # Use the account name from the position data
            account_name = position['account_name']
            
            # Create a synthetic trade entry for the current position
            trade_time = datetime.now()
            
            # Determine action based on position type
            if position['position_type'] == "LONG":
                action = "OPEN"
                contracts = position['quantity']
            else:  # SHORT
                action = "OPEN"  # Short calls are also OPEN positions
                contracts = position['quantity']
            
            # Create notes with position details
            leap_note = " (LEAP)" if position['is_leap'] else ""
            notes = f"Synced from Schwab: {position['symbol']}{leap_note}"
            if position['position_type'] == "SHORT":
                notes += " - SHORT CALL"
            
            # Create the row data
            row = {
                "Account": account_name,
                "Date": trade_time,
                "Action": action,
                "Symbol": position['underlying'].upper(),
                "Side": "Call",
                "Contracts": contracts,
                "Strike": round(position['strike'], 6),
                "Expiry": parse_date(position['expiration']) if position['expiration'] else None,
                "Premium/Buyback": 0.0,  # We don't have premium data from positions
                "Underlying": 0.0,  # We don't have current underlying price
                "Notes": notes
            }
            
            # Generate composite key
            row["Key"] = composite_key(
                row["Symbol"], 
                row["Strike"], 
                position['expiration'],
                row["Side"],
                action
            )
            
            # Determine destination file
            if args.file:
                dest = Path(args.file)
            else:
                dest = Path(f"Juice_Ledger_{account_name}.xlsx")
            
            # Append the row
            append_row(dest, row)
            synced_count += 1
            
            # Track statistics
            if position['is_leap']:
                leap_count += 1
            if position['position_type'] == "SHORT":
                short_call_count += 1
            
            position_type_desc = f"{position['position_type']} CALL"
            leap_desc = " (LEAP)" if position['is_leap'] else ""
            print(f"✅ Synced {position['underlying']} {position['strike']} {position['expiration']} - {position_type_desc} {contracts} contracts{leap_desc}")
            
        except Exception as e:
            print(f"❌ Error processing position {position.get('symbol', 'unknown')}: {e}")
            continue
    
    # Print summary
    print(f"\n📊 Sync Summary:")
    print(f"   Total positions synced: {synced_count}")
    print(f"   LEAP positions: {leap_count}")
    print(f"   Short call positions: {short_call_count}")
    print(f"✅ Successfully synced {synced_count} positions from Schwab")

def cmd_sync_trades(args):
    """Sync historical trades from Schwab to Excel ledger with execution times."""
    _ensure_schwab()
    
    print("🔄 Syncing historical trades from Schwab to Excel ledger...")
    
    # Get historical trades from Schwab
    schwab_trades = get_schwab_orders(args.account, args.days_back)
    
    if not schwab_trades:
        print("📭 No option trades found in Schwab accounts")
        return
    
    print(f"📊 Found {len(schwab_trades)} option trade executions")
    
    # Process each trade
    synced_count = 0
    buy_count = 0
    sell_count = 0
    
    for trade in schwab_trades:
        try:
            # Skip if we can't parse the symbol
            if not trade['underlying']:
                print(f"⚠️  Skipping unparseable symbol: {trade['symbol']}")
                continue
            
            # Use the account name from the trade data
            account_name = trade['account_name']
            
            # Determine action based on instruction
            if trade['instruction'] == 'SELL':
                action = "OPEN"  # Selling calls = opening short positions
                contracts = trade['quantity']
            elif trade['instruction'] == 'BUY':
                action = "CLOSE"  # Buying calls = closing short positions
                contracts = trade['quantity']
            else:
                print(f"⚠️  Skipping unknown instruction: {trade['instruction']}")
                continue
            
            # Create notes with trade details
            notes = f"Synced from Schwab: {trade['symbol']} - Executed: {trade['execution_time_str']}"
            if trade['instruction'] == 'SELL':
                notes += " - SHORT CALL"
            
            # Create the row data
            row = {
                "Account": account_name,
                "Date": trade['execution_time'],  # Use actual execution time
                "Action": action,
                "Symbol": trade['underlying'].upper(),
                "Side": "Call",
                "Contracts": contracts,
                "Strike": round(trade['strike'], 6),
                "Expiry": parse_date(trade['expiration']) if trade['expiration'] else None,
                "Premium/Buyback": trade['price'],  # Use actual execution price
                "Underlying": 0.0,  # We don't have underlying price at execution
                "Notes": notes
            }
            
            # Generate composite key
            row["Key"] = composite_key(
                row["Symbol"], 
                row["Strike"], 
                trade['expiration'],
                row["Side"],
                action
            )
            
            # Determine destination file
            if args.file:
                dest = Path(args.file)
            else:
                dest = Path(f"Juice_Ledger_{account_name}.xlsx")
            
            # Append the row
            append_row(dest, row)
            synced_count += 1
            
            # Track statistics
            if trade['instruction'] == 'BUY':
                buy_count += 1
            else:
                sell_count += 1
            
            print(f"✅ Synced {trade['underlying']} {trade['strike']} {trade['expiration']} - {action} {contracts} contracts @ ${trade['price']} (Executed: {trade['execution_time_str']})")
            
        except Exception as e:
            print(f"❌ Error processing trade {trade.get('symbol', 'unknown')}: {e}")
            continue
    
    # Print summary
    print(f"\n📊 Trade Sync Summary:")
    print(f"   Total trades synced: {synced_count}")
    print(f"   SELL trades (OPEN): {sell_count}")
    print(f"   BUY trades (CLOSE): {buy_count}")
    print(f"✅ Successfully synced {synced_count} trades from Schwab")


def cmd_batch(args):
    """Process a CSV of trades and dispatch to existing subcommands."""
    csv_path = Path(args.csv)
    if not csv_path.exists():
        raise SystemExit(f"CSV file not found: {csv_path}")

    # Resolve CSV log path and pre-load existing entries for dedupe
    csv_log_path = Path(args.csv_log) if args.csv_log else DEFAULT_CSV_LOG

    def load_ledger_keys(file_path: Path) -> set:
        """Return signatures of rows already in the ledger to avoid true duplicates.
        Signature includes key/action/account/side plus date, contracts, and price to allow repeated opens/closes at the same strike/expiry when they differ."""
        existing = set()
        if not file_path.exists():
            return existing
        _ensure_openpyxl()
        wb_data = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb_data["Ledger"]
        for r in range(2, ws.max_row + 1):
            symbol = (ws.cell(row=r, column=COLS["Symbol"]).value or "").strip()
            strike = ws.cell(row=r, column=COLS["Strike"]).value
            expiry = ws.cell(row=r, column=COLS["Expiry"]).value
            side = (ws.cell(row=r, column=COLS["Side"]).value or "Call").strip()
            action_text = (ws.cell(row=r, column=COLS["Action"]).value or "Open").strip()
            action_upper = "CLOSE" if "close" in action_text.lower() else "OPEN"
            if not symbol or strike is None or expiry is None:
                continue
            key = composite_key(symbol, strike, expiry, side, action_upper)
            account_val = str(ws.cell(row=r, column=COLS["Account"]).value or "").strip().upper()
            side_upper = side.upper()
            date_val = ws.cell(row=r, column=COLS["Date"]).value
            date_str = ""
            if hasattr(date_val, "strftime"):
                date_str = date_val.strftime("%Y-%m-%d %H:%M:%S")
            elif date_val is not None:
                date_str = str(date_val)
            contracts_val = ws.cell(row=r, column=COLS["Contracts"]).value
            price_val = ws.cell(row=r, column=COLS["Premium/Buyback"]).value
            existing.add((key.upper(), action_upper, account_val, side_upper, date_str, contracts_val, price_val))
        return existing

    def make_dedupe_key(cleaned_row: dict, command: str, action_label: str) -> tuple:
        """Create a tuple key used to detect duplicates (includes date/contracts/price)."""
        account = (cleaned_row.get("Account") or "").strip().upper()
        side_val = (cleaned_row.get("Side") or ("Call" if command in ("open", "close") else "Call")).strip()
        side = side_val.upper()
        action_upper = action_label.strip().upper()
        symbol = cleaned_row.get("Symbol") or ""
        strike_raw = cleaned_row.get("Strike")
        try:
            strike_val = float(strike_raw) if strike_raw else None
        except ValueError:
            strike_val = None
        expiry_val = cleaned_row.get("Expiry") or None
        key_val = composite_key(symbol, strike_val, expiry_val, side_val, action_upper)
        date_part = (cleaned_row.get("Date") or "").strip()
        time_part = (cleaned_row.get("Time") or "").strip()
        datetime_part = f"{date_part} {time_part}".strip()
        contracts_val = None
        try:
            contracts_val = int(cleaned_row["Contracts"]) if cleaned_row.get("Contracts") else None
        except (ValueError, TypeError, KeyError):
            contracts_val = cleaned_row.get("Contracts")
        price_val = None
        price_field = "Premium" if command == "open" else ("Buyback" if command == "close" else "Premium")
        try:
            price_val = float(cleaned_row[price_field]) if cleaned_row.get(price_field) else None
        except (ValueError, TypeError, KeyError):
            price_val = cleaned_row.get(price_field)
        return (key_val.upper(), action_upper, account, side, datetime_part, contracts_val, price_val)

    ledger_key_cache: dict[str, set] = {}
    ledger_open_counts_cache: dict[str, dict[tuple[str, str], int]] = {}
    pending_close_counts: defaultdict[tuple[str, str], int] = defaultdict(int)

    def keys_for_account(account_name: str, file_arg: Optional[str]) -> set:
        key = account_name.strip()
        if key not in ledger_key_cache:
            path = ensure_account_workbook(account_name, file_arg or "")
            ledger_key_cache[key] = load_ledger_keys(path)
        return ledger_key_cache[key]

    def open_counts_for_account(account_name: str, file_arg: Optional[str]) -> dict[tuple[str, str], int]:
        key = account_name.strip()
        if key not in ledger_open_counts_cache:
            path = ensure_account_workbook(account_name, file_arg or "")
            counts: dict[tuple[str, str], int] = defaultdict(int)
            if path.exists():
                wb_data = openpyxl.load_workbook(path, data_only=True)
                ws = wb_data["Ledger"]
                for r in range(2, ws.max_row + 1):
                    acc = (ws.cell(row=r, column=COLS["Account"]).value or "").strip().upper()
                    symbol = (ws.cell(row=r, column=COLS["Symbol"]).value or "").strip().upper()
                    strike = ws.cell(row=r, column=COLS["Strike"]).value
                    expiry = ws.cell(row=r, column=COLS["Expiry"]).value
                    raw_side = ws.cell(row=r, column=COLS["Side"]).value
                    side_val = (raw_side if raw_side not in (None, "") else "Call")
                    side = str(side_val).strip().capitalize() or "Call"
                    action_text = (ws.cell(row=r, column=COLS["Action"]).value or "").strip().upper()
                    action = "CLOSE" if "CLOSE" in action_text else "OPEN"
                    contracts = ws.cell(row=r, column=COLS["Contracts"]).value
                    if not acc or not symbol or strike is None or expiry is None or contracts is None:
                        continue
                    try:
                        strike_val = float(strike)
                    except (TypeError, ValueError):
                        strike_val = strike
                    expiry_val = expiry
                    open_key = composite_key(symbol, strike_val, expiry_val, side, "OPEN").upper()
                    key_tuple = (acc, open_key)
                    if action == "OPEN":
                        counts[key_tuple] += int(contracts)
                    else:
                        counts[key_tuple] -= int(contracts)
            ledger_open_counts_cache[key] = counts
        return ledger_open_counts_cache[key]

    def build_batch_rows():
        rows = []
        order_map = {"open": 0, "close": 1, "jl": 2}
        with csv_path.open(newline="", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            for row_num, row in enumerate(reader, start=2):
                if not row:
                    continue

                cleaned = {k: (v.strip() if isinstance(v, str) else v) for k, v in row.items()}
                command_raw = cleaned.get("Command", "") or ""
                command = command_raw.strip().lower()

                if not command or command.startswith("#"):
                    continue

                try:
                    trade_dt = parse_datetime(cleaned.get("Date"), cleaned.get("Time"))
                except SystemExit:
                    trade_dt = datetime.min

                raw_action = (cleaned.get("Action") or "").strip()
                if raw_action:
                    action_lower = raw_action.lower()
                    if "close" in action_lower:
                        action_label = "Close"
                    elif "open" in action_lower:
                        action_label = "Open"
                    else:
                        action_label = "Close" if command == "close" else "Open"
                else:
                    action_label = "Close" if command == "close" else "Open"

                base_key = batch_base_key(cleaned)
                rows.append((trade_dt, order_map.get(command, 3), row_num, command, command_raw, cleaned, action_label, base_key))
        return sorted(rows, key=lambda entry: (entry[0], entry[1], entry[2]))

    rows = build_batch_rows()
    cache_requirements: dict[tuple[str, str], datetime] = {}
    if args.auto_price:
        for trade_dt, _, _, command, _, cleaned, _, _ in rows:
            symbol_val = (cleaned.get("Symbol") or "").upper()
            if not symbol_val or not isinstance(trade_dt, datetime) or trade_dt.year < 1900:
                continue
            if command not in {"open", "close"}:
                continue
            price_key = "Underlying" if command == "open" else "UnderlyingClose"
            if cleaned.get(price_key):
                continue
            month_key = f"{trade_dt.year}-{trade_dt.month:02d}"
            cache_requirements.setdefault((symbol_val, month_key), trade_dt)
    missing_combos: dict[tuple[str, str], datetime] = {}
    if args.auto_price:
        for (symbol_val, month_key), trade_dt in cache_requirements.items():
            cache_file = _alpha_cache_path(symbol_val, month_key)
            if not cache_file.exists():
                missing_combos[(symbol_val, month_key)] = trade_dt
    if missing_combos:
        print("Alpha Vantage cache missing for:")
        for sym, month in sorted(missing_combos):
            print(f"  {sym} {month}")
    if args.auto_price and missing_combos:
        print("Fetching missing Alpha Vantage data to populate cache...")
        for (symbol_val, month_key), trade_dt in sorted(missing_combos.items()):
            eastern_dt = convert_central_to_eastern(trade_dt)
            fetched = fetch_alpha_vantage_time_series(symbol_val, eastern_dt)
            if fetched is None:
                print(f"⚠️ Failed to populate Alpha Vantage cache for {symbol_val} {month_key}")
            else:
                print(f"✅ Cached Alpha Vantage data for {symbol_val} {month_key}")
    log_ledger_positions([Path(args.file)] if args.file else [Path("Juice_Ledger_Travis.xlsx"), Path("Juice_Ledger_Christie.xlsx")])
    summary = summarize_batch_rows(rows)
    log_batch_stats(summary, Path(csv_path))
    log_batch_keys(rows)
    ledger_paths = [Path(args.file)] if args.file else [Path("Juice_Ledger_Travis.xlsx"), Path("Juice_Ledger_Christie.xlsx")]
    ledger_key_tuples = collect_ledger_key_tuples(ledger_paths)
    if ledger_key_tuples:
        print("Ledger keys available:")
        for key, action, account, side in sorted(ledger_key_tuples):
            print(f"  {key} ({action}, {account}, {side})")
    ledger_base_keys = {base_from_composite(key) for key, *_ in ledger_key_tuples}
    matching_opens = summary["opens"] & ledger_base_keys
    matching_closes = summary["closes"] & ledger_base_keys
    print("Batch ledger matches:")
    for key in sorted(matching_opens):
        print(f"  Ledger has open key also seen in batch open: {key}")
    for key in sorted(matching_closes):
        print(f"  Ledger has key also seen in batch close: {key}")

    if args.auto_price:
        for trade_dt, _, row_num, command, _, cleaned, _, _ in rows:
            if command not in {"open", "close"}:
                continue
            price_key = "Underlying" if command == "open" else "UnderlyingClose"
            if cleaned.get(price_key):
                continue
            price = get_cached_alpha_price(cleaned.get("Symbol"), trade_dt)
            if price is not None:
                cleaned[price_key] = str(price)
                print(f"Row {row_num}: used cached price for {cleaned.get('Symbol')} -> {price_key}={price}")

    for trade_dt, _, row_num, command, command_raw, cleaned, action_label, base_key in rows:
        dedupe_key = make_dedupe_key(cleaned, command, action_label)
        print(f"Row {row_num}: CSV action '{command_raw}' -> Key={dedupe_key[0]} (Action={dedupe_key[1]}, Side={dedupe_key[3]})")
        ledger_account = cleaned.get("Account") or "Unknown"
        ledger_keys = keys_for_account(ledger_account, args.file)
        if dedupe_key in ledger_keys:
            print(f"⚠️  Skipping row {row_num}: already exists (Key={dedupe_key[0]}, Action={dedupe_key[1]}, Account={dedupe_key[2]}, Side={dedupe_key[3]})")
            continue

        # Track remaining open contracts for this account/key to allow multiple partial closes
        acc_upper = ledger_account.strip().upper()
        side_for_key = (cleaned.get("Side") or ("Call" if command in ("open", "close") else "Call")).strip().capitalize() or "Call"
        strike_raw = cleaned.get("Strike")
        try:
            strike_for_key = float(strike_raw) if strike_raw else None
        except ValueError:
            strike_for_key = strike_raw
        expiry_for_key = cleaned.get("Expiry") or None
        open_key_upper = composite_key(cleaned.get("Symbol") or "", strike_for_key, expiry_for_key, side_for_key, "OPEN").upper()
        open_counts = open_counts_for_account(ledger_account, args.file)
        key_tuple = (acc_upper, open_key_upper)
        if command == "close":
            available_now = open_counts.get(key_tuple, 0) - pending_close_counts[key_tuple]
            contracts_requested = abs(int(cleaned["Contracts"])) if cleaned.get("Contracts") else 0
            if available_now <= 0:
                print(f"⚠️  Skipping row {row_num}: no open contracts remain for {open_key_upper} (Account={acc_upper})")
                continue
            if contracts_requested > available_now:
                print(f"⚠️  Row {row_num}: requested close {contracts_requested} but only {available_now} remain open for {open_key_upper}. Closing available {available_now}.")
                cleaned["Contracts"] = available_now
                contracts_requested = available_now

        try:
            if command == "open":
                ns = argparse.Namespace(
                    file=args.file,
                    account=cleaned.get("Account"),
                    symbol=cleaned.get("Symbol"),
                    contracts=abs(int(cleaned["Contracts"])) if cleaned.get("Contracts") else 0,
                    date=cleaned.get("Date") or None,
                    time=cleaned.get("Time") or None,
                    premium=float(cleaned["Premium"]) if cleaned.get("Premium") else 0.0,
                    strike=float(cleaned["Strike"]) if cleaned.get("Strike") else None,
                    side=(cleaned.get("Side") or "Call"),
                    underlying=float(cleaned["Underlying"]) if cleaned.get("Underlying") else None,
                    expiry=cleaned.get("Expiry") or None,
                    auto_price=bool(args.auto_price and not cleaned.get("Underlying")),
                    notes=cleaned.get("Notes") or "",
                    csv_log=str(csv_log_path),
                )
                cmd_open(ns)
                print(f"Processed row {row_num}: OPEN {ns.account} {ns.symbol} {ns.strike} {ns.expiry}")
                ledger_keys.add(dedupe_key)
                if ns.contracts and open_key_upper:
                    open_counts[key_tuple] = open_counts.get(key_tuple, 0) + ns.contracts
            elif command == "close":
                ns = argparse.Namespace(
                    file=args.file,
                    account=cleaned.get("Account"),
                    symbol=cleaned.get("Symbol"),
                    contracts=abs(int(cleaned["Contracts"])) if cleaned.get("Contracts") else 0,
                    date=cleaned.get("Date") or None,
                    time=cleaned.get("Time") or None,
                    buyback=float(cleaned["Buyback"]) if cleaned.get("Buyback") else 0.0,
                    strike=float(cleaned["Strike"]) if cleaned.get("Strike") else None,
                    side=(cleaned.get("Side") or "Call"),
                    underlying_close=float(cleaned["UnderlyingClose"]) if cleaned.get("UnderlyingClose") else None,
                    expiry=cleaned.get("Expiry") or None,
                    auto_price=bool(args.auto_price and not cleaned.get("UnderlyingClose")),
                    notes=cleaned.get("Notes") or "",
                    csv_log=str(csv_log_path),
                )
                cmd_close(ns)
                print(f"Processed row {row_num}: CLOSE {ns.account} {ns.symbol} {ns.strike} {ns.expiry}")
                pending_close_counts[key_tuple] += ns.contracts
            elif command == "jl":
                ns = argparse.Namespace(
                    ticker=cleaned.get("Symbol"),
                    action=(cleaned.get("Action") or "Open"),
                    side=(cleaned.get("Side") or "Call"),
                    strike=float(cleaned["Strike"]) if cleaned.get("Strike") else None,
                    expiry=cleaned.get("Expiry") or None,
                    underlying=float(cleaned["Underlying"]) if cleaned.get("Underlying") else None,
                    premium=float(cleaned["Premium"]) if cleaned.get("Premium") else 0.0,
                    contracts=abs(int(cleaned["Contracts"])) if cleaned.get("Contracts") else 1,
                    date=cleaned.get("Date") or None,
                    time=cleaned.get("Time") or None,
                    auto_price=bool(args.auto_price and not cleaned.get("Underlying")),
                    account=cleaned.get("Account") or "Juice Lever",
                    note=cleaned.get("Notes") or "",
                    csv_log=str(csv_log_path),
                )
                cmd_jl(ns)
                print(f"Processed row {row_num}: JL {ns.action} {ns.side} {ns.ticker} {ns.strike} {ns.expiry}")
                processed_keys.add(dedupe_key)
            else:
                    print(f"⚠️  Skipping row {row_num}: unknown command '{command_raw}'")
        except Exception as e:
            print(f"❌ Error processing row {row_num}: {e}")
            print(f"   Row data: {cleaned}")

def cmd_prompt():
    """
    Interactive wizard that keeps defaults between entries so you can log several rows quickly.
    Hit Enter to reuse the last value in brackets. Type 'q' when asked to add another to exit.
    """
    print("\n🪄 CFM Trade Entry Wizard (session mode)")
    print("Press Enter to reuse the default in [brackets]; change any field as needed.\n")

    session_defaults = {
        "account": "Travis",
        "action": "Open",
        "symbol": "",
        "side": "Call",
        "contracts": 1,
        "strike": None,
        "expiry": "",
        "notes": "",
        "file": "",
    }
    entry_num = 1
    while True:
        print(f"\n--- Entry #{entry_num} ---")
        account = prompt_choice("Account", ["Travis", "Christie"], default=session_defaults["account"])
        action = prompt_choice("Action", ["Open", "Close"], default=session_defaults["action"])
        symbol_default = session_defaults["symbol"]
        symbol_input = prompt_value("Ticker (e.g., NVDA)", default=symbol_default or None, required=not bool(symbol_default))
        symbol = symbol_input.upper()
        side = prompt_choice("Side", ["Call", "Put"], default=session_defaults["side"])
        contracts = prompt_number("Contracts", int, default=session_defaults["contracts"], required=True)
        strike = prompt_number("Strike price", float, default=session_defaults["strike"], required=True)
        expiry = prompt_value("Expiry (YYYY-MM-DD, blank allowed)", default=session_defaults["expiry"], required=False)
        date_str = prompt_value("Trade date (YYYY-MM-DD, blank=today)", default="", required=False)
        time_str = prompt_value("Time (HH:MM, blank=now)", default="", required=False)
        notes = prompt_value("Notes (optional)", default=session_defaults["notes"], required=False)
        file_override = prompt_value("Ledger file override (blank=auto)", default=session_defaults["file"], required=False)

        if action.lower() == "open":
            premium = prompt_number("Premium per contract", float, required=True)
            underlying = prompt_number("Underlying price (blank to auto-fetch)", float, required=False, default=None)
            auto_price = underlying is None
            args = argparse.Namespace(
                file=file_override or None,
                account=account,
                symbol=symbol,
                contracts=contracts,
                date=date_str or None,
                time=time_str or None,
                premium=premium,
                strike=strike,
                side=side,
                underlying=underlying,
                expiry=expiry or None,
                auto_price=auto_price,
                notes=notes or "",
                csv_log=None,
            )
            cmd_open(args)
        else:
            buyback = prompt_number("Buyback per contract", float, required=True)
            underlying_close = prompt_number("Underlying at close (blank to auto-fetch)", float, required=False, default=None)
            auto_price = underlying_close is None
            expiry_close = expiry or session_defaults["expiry"] or prompt_value("Expiry is required to close (YYYY-MM-DD)", required=True)
            args = argparse.Namespace(
                file=file_override or None,
                account=account,
                symbol=symbol,
                contracts=contracts,
                date=date_str or None,
                time=time_str or None,
                buyback=buyback,
                strike=strike,
                side=side,
                underlying_close=underlying_close,
                expiry=expiry_close,
                auto_price=auto_price,
                notes=notes or "",
                csv_log=None,
            )
            cmd_close(args)

        # Update defaults from this entry for the next loop
        session_defaults.update({
            "account": account,
            "action": action,
            "symbol": symbol,
            "side": side,
            "contracts": contracts,
            "strike": strike,
            "expiry": expiry,
            "notes": notes,
            "file": file_override,
        })
        entry_num += 1

        cont = input("Add another entry? [Y/n]: ").strip().lower()
        if cont and cont.startswith("n"):
            break

def build():
    p = argparse.ArgumentParser(
        prog="cfm-ledger",
        description="Append-only JUICE ledger (supports partial closes via composite key). If account file is missing, clones Juice_Ledger.xlsx or creates fresh."
    )
    sub = p.add_subparsers(dest="cmd", required=True)
    po = sub.add_parser("open", help="Open short call (append an OPEN lot)")
    po.add_argument("--file", help="Excel file (defaults to Juice_Ledger_{Account}.xlsx)")
    po.add_argument("--account", required=True, choices=["Travis","Christie"])
    po.add_argument("--symbol", required=True)
    po.add_argument("--contracts", required=True, type=int)
    po.add_argument("--date", help="YYYY-MM-DD or MM/DD/YYYY (defaults to today)")
    po.add_argument("--time", help="HH:MM (defaults to current time)")
    po.add_argument("--premium", type=float, required=True, help="Premium per contract")
    po.add_argument("--strike", type=float, required=True, help="Strike price")
    po.add_argument("--side", choices=["Call","Put"], default="Call", help="Option side to record (default: Call)")
    po.add_argument("--underlying", type=float, help="Underlying price (auto-fetched if --auto-price used)")
    po.add_argument("--expiry", help="YYYY-MM-DD or MM/DD/YYYY")
    po.add_argument("--auto-price", action="store_true", help="Auto-fetch underlying price from Yahoo Finance")
    po.add_argument("--condition", choices=["GREEN","YELLOW","RED"], help="Stock condition at entry (optional)")
    po.add_argument("--notes")
    po.add_argument("--csv-log", help="CSV log file for cash events (defaults to cfm_trades_log.csv)")
    pc = sub.add_parser("close", help="Close short call (append a CLOSE lot)")
    pc.add_argument("--file", help="Excel file (defaults to Juice_Ledger_{Account}.xlsx)")
    pc.add_argument("--account", required=True, choices=["Travis","Christie"])
    pc.add_argument("--symbol", required=True)
    pc.add_argument("--contracts", required=True, type=int, help="Number of contracts to close now")
    pc.add_argument("--date", help="YYYY-MM-DD or MM/DD/YYYY (defaults to today)")
    pc.add_argument("--time", help="HH:MM (defaults to current time)")
    pc.add_argument("--buyback", type=float, required=True, help="Buyback premium per contract")
    pc.add_argument("--strike", type=float, required=True, help="Strike to match the open lot(s)")
    pc.add_argument("--side", choices=["Call","Put"], default="Call", help="Option side to record (default: Call)")
    pc.add_argument("--underlying-close", type=float, help="Underlying at close (auto-fetched if --auto-price used)")
    pc.add_argument("--expiry", required=True, help="Expiry to match the open lot(s)")
    pc.add_argument("--auto-price", action="store_true", help="Auto-fetch underlying price from Yahoo Finance")
    pc.add_argument("--condition", choices=["GREEN","YELLOW","RED"], help="Stock condition at close (optional)")
    pc.add_argument("--notes")
    pc.add_argument("--csv-log", help="CSV log file for cash events (defaults to cfm_trades_log.csv)")
    pj = sub.add_parser("jl", help="Log Juice Lever cash events to the CSV journal")
    pj.add_argument("--ticker", required=True, help="Underlying ticker (e.g., XLF)")
    pj.add_argument("--action", required=True, choices=["Open","Close"], help="Open for short entry (credit) or Close for buyback (debit)")
    pj.add_argument("--side", choices=["Call","Put"], default="Call", help="Option side to record (Call or Put)")
    pj.add_argument("--strike", type=float, required=True, help="Strike price of the JL leg")
    pj.add_argument("--expiry", help="Expiry date (YYYY-MM-DD or MM/DD/YYYY)")
    pj.add_argument("--underlying", type=float, help="Underlying price for extrinsic calculation")
    pj.add_argument("--premium", type=float, required=True, help="Premium amount per JL contract")
    pj.add_argument("--contracts", type=int, default=1, help="Number of JL contracts (default: 1)")
    pj.add_argument("--date", help="YYYY-MM-DD or MM/DD/YYYY (defaults to today)")
    pj.add_argument("--time", help="HH:MM (defaults to current time)")
    pj.add_argument("--auto-price", action="store_true", help="Auto-fetch underlying price from Yahoo Finance")
    pj.add_argument("--account", default="Juice Lever", help="Account label to use for this log row")
    pj.add_argument("--note", help="Optional free-text comment")
    pj.add_argument("--csv-log", help="CSV log file for JL events (defaults to cfm_trades_log.csv)")
    pj_summary = sub.add_parser("jl-summary", help="Show weekly JL net juice totals from the CSV log")
    pj_summary.add_argument("--csv-log", help="CSV log file for JL events (defaults to cfm_trades_log.csv)")
    ph = sub.add_parser("highlight", help="Highlight rows with open positions")
    ph.add_argument("--file", help="Excel file (defaults to Juice_Ledger_{Account}.xlsx or auto-detect)")
    ph.add_argument("--account", choices=["Travis","Christie"], help="Account to highlight (optional, highlights all if not specified)")
    ps = sub.add_parser("sync", help="Sync positions from Schwab to Excel ledger")
    ps.add_argument("--file", help="Excel file (defaults to Juice_Ledger_{Account}.xlsx)")
    ps.add_argument("--account", choices=["Travis","Christie"], help="Account to sync to (defaults to Travis)")
    pt = sub.add_parser("sync-trades", help="Sync historical trades from Schwab with execution times")
    pt.add_argument("--file", help="Excel file (defaults to Juice_Ledger_{Account}.xlsx)")
    pt.add_argument("--account", choices=["Travis","Christie"], help="Account to sync to (defaults to Travis)")
    pt.add_argument("--days-back", type=int, default=30, help="Number of days to look back for trades (default: 30)")
    sub.add_parser("prompt", help="Interactive wizard for quickly logging a single trade")
    pb = sub.add_parser("batch", help="Process multiple trades from a CSV file")
    pb.add_argument("--csv", required=True, help="CSV file with trades to process")
    pb.add_argument("--file", help="Excel file (defaults to Juice_Ledger_{Account}.xlsx)")
    pb.add_argument("--csv-log", help="CSV log file for cash events (defaults to cfm_trades_log.csv)")
    pb.add_argument("--auto-price", action="store_true", help="Allow per-row auto price fetch when underlying is blank")
    return p

def main():
    parser = build()
    args = parser.parse_args()
    if args.cmd == "open":
        cmd_open(args)
    elif args.cmd == "close":
        cmd_close(args)
    elif args.cmd == "highlight":
        cmd_highlight(args)
    elif args.cmd == "sync":
        cmd_sync(args)
    elif args.cmd == "sync-trades":
        cmd_sync_trades(args)
    elif args.cmd == "jl":
        cmd_jl(args)
    elif args.cmd == "jl-summary":
        cmd_jl_summary(args)
    elif args.cmd == "prompt":
        cmd_prompt()
    elif args.cmd == "batch":
        cmd_batch(args)

if __name__ == "__main__":
    main()
