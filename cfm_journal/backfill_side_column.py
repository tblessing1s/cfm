"""
Backfill script that adds a Side column to the Juice Ledger workbooks and ensures the composite key
includes the Side suffix.

Usage:
  python backfill_side_column.py
"""

from pathlib import Path
from datetime import datetime
from typing import List

from openpyxl import load_workbook


def ensure_column_exists(ws, header: List[str], name: str, after: str = None) -> int:
    """
    Ensure column exists in worksheet with header row. Returns 1-based column index.
    Inserts a new column after 'after' if provided; otherwise appends at end.
    """
    if name in header:
        return header.index(name) + 1

    if after and after in header:
        insert_idx = header.index(after) + 2
    else:
        insert_idx = ws.max_column + 1

    ws.insert_cols(insert_idx)
    ws.cell(row=1, column=insert_idx).value = name
    header.insert(insert_idx - 1, name)
    return insert_idx


def move_column_to_end(ws, column_idx: int) -> int:
    """Move the specified column to the far right and return its new index."""
    values = [ws.cell(row=r, column=column_idx).value for r in range(1, ws.max_row + 1)]
    header_name = values[0]
    ws.delete_cols(column_idx)
    insert_idx = ws.max_column + 1
    ws.insert_cols(insert_idx)
    for row_idx, value in enumerate(values, start=1):
        ws.cell(row=row_idx, column=insert_idx).value = value
    ws.cell(row=1, column=insert_idx).value = header_name
    return insert_idx


def composite_key(symbol: str, strike, expiry, side: str, action: str) -> str:
    """
    Mirror of cfm_ledger_autotemplate.composite_key so the key aligns with the ledger logic.
    """
    from cfm_ledger_autotemplate import parse_date

    exp = ""
    if expiry:
        try:
            exp = parse_date(expiry).strftime("%Y-%m-%d")
        except (SystemExit, Exception):
            exp = str(expiry)

    s_strike = ""
    if strike is not None:
        try:
            strike_float = float(strike)
            if strike_float.is_integer():
                s_strike = str(int(strike_float))
            else:
                s_strike = f"{strike_float:.6f}".rstrip("0").rstrip(".")
        except (ValueError, TypeError):
            s_strike = str(strike)

    side_part = (side or "Call").upper()
    action_part = (action or "Open").upper()
    return f"{(symbol or '').upper()}|{s_strike}|{exp}|{side_part}|{action_part}"


def backfill(path: Path):
    wb = load_workbook(path)
    ws = wb["Ledger"]
    header = [ws.cell(row=1, column=col).value or "" for col in range(1, ws.max_column + 1)]

    symbol_col = header.index("Symbol") + 1 if "Symbol" in header else None
    strike_col = header.index("Strike") + 1 if "Strike" in header else None
    expiry_col = header.index("Expiry") + 1 if "Expiry" in header else None
    key_col = header.index("Key") + 1 if "Key" in header else None

    if symbol_col is None or strike_col is None or expiry_col is None:
        raise SystemExit(f"Workbook {path.name} is missing required columns.")

    side_col = header.index("Side") + 1 if "Side" in header else None
    if side_col:
        if side_col != len(header):
            side_col = move_column_to_end(ws, side_col)
        header = [ws.cell(row=1, column=col).value or "" for col in range(1, ws.max_column + 1)]
    else:
        side_col = ensure_column_exists(ws, header, "Side")
        header = [ws.cell(row=1, column=col).value or "" for col in range(1, ws.max_column + 1)]
        if side_col != len(header):
            side_col = move_column_to_end(ws, side_col)
            header = [ws.cell(row=1, column=col).value or "" for col in range(1, ws.max_column + 1)]
    if "Key" in header:
        key_col = header.index("Key") + 1
    else:
        key_col = ensure_column_exists(ws, header, "Key")
        header = [ws.cell(row=1, column=col).value or "" for col in range(1, ws.max_column + 1)]

    symbol_col = header.index("Symbol") + 1
    strike_col = header.index("Strike") + 1
    expiry_col = header.index("Expiry") + 1
    action_col = header.index("Action") + 1 if "Action" in header else None
    side_col = header.index("Side") + 1
    key_col = header.index("Key") + 1

    for row in range(2, ws.max_row + 1):
        symbol = ws.cell(row=row, column=symbol_col).value or ""
        strike = ws.cell(row=row, column=strike_col).value
        expiry = ws.cell(row=row, column=expiry_col).value
        expiry_value = expiry
        if isinstance(expiry, datetime):
            expiry_value = expiry.strftime("%Y-%m-%d")

        raw_side = ws.cell(row=row, column=side_col).value
        side = raw_side.strip() if isinstance(raw_side, str) and raw_side.strip() else "Call"
        ws.cell(row=row, column=side_col).value = side

        action_value = "Open"
        if action_col:
            raw_action = ws.cell(row=row, column=action_col).value
            if isinstance(raw_action, str) and raw_action.strip():
                action_value = raw_action.strip()
            else:
                action_value = "Open"

        new_key = composite_key(symbol, strike, expiry_value, side, action_value)
        ws.cell(row=row, column=key_col).value = new_key

    wb.save(path)
    print(f"Backfilled {path.name}: added Side column and refreshed Key entries.")


def main():
    ledger_files = [
        Path("Juice_Ledger_Christie.xlsx"),
        Path("Juice_Ledger_Travis.xlsx"),
    ]

    for ledger in ledger_files:
        if not ledger.exists():
            print(f"Missing {ledger}, skipping.")
            continue
        backfill(ledger)


if __name__ == "__main__":
    main()
