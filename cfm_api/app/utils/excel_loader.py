"""Helpers for loading trade data from per-account ledgers."""
from __future__ import annotations

from dataclasses import dataclass
import csv
from pathlib import Path
from typing import Any, Dict, List, Union
import shutil
import json
import os
import requests

import numpy as np
import pandas as pd
import openpyxl

BASE_DIR = Path(__file__).resolve().parents[3]
JOURNAL_ROOT = BASE_DIR / "cfm_journal"
DATA_DIR = JOURNAL_ROOT / "data"
ALPHA_CACHE_DIR = JOURNAL_ROOT / "alpha_cache"
DATA_DIR.mkdir(parents=True, exist_ok=True)
LEDGER_PATTERN = "Juice_Ledger*.xlsx"
UI_TRADES_FILE = DATA_DIR / "ui_trades.csv"


@dataclass(frozen=True)
class _AccountInfo:
    name: str
    label: str
    path: Path


def _friendly_label(stem: str) -> str:
    normalized = stem.lower()
    prefix = "juice_ledger_"

    if normalized.startswith(prefix):
        label = stem[len(prefix):]
    elif normalized == "juice_ledger":
        label = ""
    else:
        label = stem

    label = label.replace("_", " ").strip()
    if not label:
        return "CFM Combined"
    return label.title()


def _discover_accounts() -> List[_AccountInfo]:
    roots = [DATA_DIR, JOURNAL_ROOT]
    collected: Dict[str, Path] = {}
    for root in roots:
        if not root.exists():
            continue
        for path in sorted(root.glob(LEDGER_PATTERN)):
            if path.stem.lower() in {"juice_ledger"}:
                continue
            stem = path.stem.lower()
            # Prefer DATA_DIR; avoid moving files during requests to prevent hangs.
            target = DATA_DIR / path.name
            if stem not in collected:
                collected[stem] = target if target.exists() else path

    if not collected:
        raise FileNotFoundError(f"No ledger files found under {DATA_DIR} or {JOURNAL_ROOT}")

    return [_AccountInfo(name=path.stem, label=_friendly_label(path.stem), path=path) for path in collected.values()]


def _resolve_account(account: str | None) -> _AccountInfo:
    descriptors = _discover_accounts()
    lookup: Dict[str, _AccountInfo] = {}

    for descriptor in descriptors:
        lookup[descriptor.name.lower()] = descriptor
        lookup[descriptor.label.lower()] = descriptor
        if descriptor.label.lower() == "cfm combined":
            for alias in ("combined", "all"):
                lookup[alias] = descriptor

    if account:
        normalized = account.strip().lower()
        candidate = lookup.get(normalized)
        if candidate:
            return candidate
        raise ValueError(f"No ledger matches account '{account}'")

    return descriptors[0]


def _normalize_columns(columns: List[str]) -> List[str]:
    return [col.strip().lower().replace(" ", "_") for col in columns]


def _composite_key(ticker: str | None, strike: Any, expiry: Any, side: str | None, action: str | None) -> str | None:
    if not ticker:
        return None
    exp_str = ""
    if expiry:
        try:
            exp_str = pd.to_datetime(expiry).strftime("%Y-%m-%d")
        except Exception:
            exp_str = str(expiry).split(" ")[0]
    try:
        strike_val = float(strike) if strike is not None else None
    except Exception:
        strike_val = None
    strike_str = ""
    if strike_val is not None:
        strike_str = str(int(strike_val)) if strike_val.is_integer() else f"{strike_val:.6f}".rstrip("0").rstrip(".")
    side_part = (side or "CALL").upper()
    action_part = (action or "OPEN").upper()
    return f"{str(ticker).upper()}|{strike_str}|{exp_str}|{side_part}|{action_part}"


def _serialize_date(value: pd.Timestamp | None) -> str | None:
    if value is pd.NaT or value is None:
        return None
    return value.date().isoformat()


def _load_trades(account: str | None = None) -> pd.DataFrame:
    account_descriptor = _resolve_account(account)
    df = pd.read_excel(account_descriptor.path, engine="openpyxl")
    df.columns = _normalize_columns(df.columns.tolist())

    if "date" not in df.columns:
        raise KeyError("Expected a 'date' column in trades Excel file")

    df = df.copy()
    df["date"] = pd.to_datetime(df["date"], errors="coerce")

    ui_trades = _load_ui_trades(account_descriptor.name)
    if not ui_trades.empty:
        df = pd.concat([df, ui_trades], ignore_index=True, sort=False)

    return df


def get_available_accounts() -> List[Dict[str, str]]:
    return [
        {"name": descriptor.name, "label": descriptor.label}
        for descriptor in _discover_accounts()
    ]


def get_all_trades(account: str | None = None) -> pd.DataFrame:
    df = _load_trades(account).copy()
    df = _ensure_core_columns(df)
    df = _normalize_missing(df)
    if not df.empty:
        df["date"] = pd.to_datetime(df["date"], errors="coerce")
    return df.drop(columns=["notes"], errors="ignore")


def _ensure_core_columns(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return df

    column_defaults: Dict[str, Union[float, str, None]] = {
        "ticker": "",
        "strategy": "CFM",
        "juice": 0.0,
        "basis": 0.0,
        "premium_in": None,
        "premium_out": None,
        "dte": None,
        "itm": None,
    }

    for column, default in column_defaults.items():
        if column not in df.columns:
            df[column] = default

    df["juice"] = pd.to_numeric(df["juice"], errors="coerce").fillna(0.0)
    df["basis"] = pd.to_numeric(df["basis"], errors="coerce").fillna(0.0)
    df["premium_in"] = pd.to_numeric(df["premium_in"], errors="coerce")
    df["premium_out"] = pd.to_numeric(df["premium_out"], errors="coerce")
    df["dte"] = pd.to_numeric(df["dte"], errors="coerce").astype("Int64")

    return df


def _normalize_missing(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return df

    return df.replace({np.nan: None})


def _compute_juice_fields(record: Dict[str, Any]) -> Dict[str, Any]:
    """Compute juice per contract and signed juice fields similar to the Excel formulas."""
    action = str(record.get("action") or "Open").upper()
    side = str(record.get("side") or "Call").strip().lower()
    strike = record.get("strike")
    underlying = record.get("underlying")
    premium = record.get("premium_buyback")
    contracts = record.get("contracts")

    def to_float(val):
        try:
            return float(val)
        except (TypeError, ValueError):
            return None

    strike_f = to_float(strike)
    underlying_f = to_float(underlying)
    premium_f = to_float(premium)
    contracts_i = None
    try:
        contracts_i = int(contracts) if contracts is not None else None
    except (TypeError, ValueError):
        contracts_i = None

    juice_per_contract = None

    if premium_f is not None:
        if strike_f is not None and underlying_f is not None:
            intrinsic = max(0, strike_f - underlying_f) if "put" in side else max(0, underlying_f - strike_f)
            extrinsic = premium_f - intrinsic
            if action == "CLOSE":
                juice_per_contract = abs(extrinsic) if extrinsic < 0 else -extrinsic
            else:
                juice_per_contract = extrinsic
        else:
            if action == "CLOSE":
                juice_per_contract = abs(premium_f) if premium_f < 0 else -premium_f
            else:
                juice_per_contract = premium_f

    signed_juice_dollars = None
    signed_juice_per_100 = None

    if juice_per_contract is not None and contracts_i is not None:
        signed_juice_dollars = round(juice_per_contract * contracts_i, 2)
        signed_juice_per_100 = round(signed_juice_dollars * 100, 2)

    def r2(val):
        return round(val, 2) if isinstance(val, (int, float)) else val

    return {
        "juice_per_contract": r2(juice_per_contract) if juice_per_contract is not None else None,
        "signed_juice_dollars": r2(signed_juice_dollars) if signed_juice_dollars is not None else None,
        "signed_juice_per_100": r2(signed_juice_per_100) if signed_juice_per_100 is not None else None,
    }


def resolve_account_name(account: str | None) -> str:
    """Return the canonical account name used on disk, validating the input."""
    descriptor = _resolve_account(account)
    return descriptor.name


def resolve_account(account: str | None) -> _AccountInfo:
    """Return the full account descriptor including the ledger path."""
    return _resolve_account(account)


def _load_ui_trades(account_name: str | None) -> pd.DataFrame:
    """Load UI-submitted trades stored in a lightweight CSV."""
    if not UI_TRADES_FILE.exists():
        return pd.DataFrame()

    df = pd.read_csv(UI_TRADES_FILE)
    if df.empty:
        return df

    df.columns = _normalize_columns(df.columns.tolist())
    if account_name and "account" in df.columns:
        df = df[df["account"].astype(str).str.lower() == account_name.lower()]

    df = df.copy()
    df["date"] = pd.to_datetime(df["date"], errors="coerce")
    return df


def append_ui_trades(rows: List[Dict[str, Any]]) -> int:
    """Append UI-submitted trades to a simple CSV store under cfm_journal."""
    if not rows:
        return 0

    UI_TRADES_FILE.parent.mkdir(parents=True, exist_ok=True)
    fieldnames = [
        "account",
        "date",
        "ticker",
        "strategy",
        "premium_in",
        "premium_out",
        "juice",
        "basis",
        "dte",
        "itm",
    ]
    exists = UI_TRADES_FILE.exists()

    with UI_TRADES_FILE.open("a", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        if not exists:
            writer.writeheader()

        for row in rows:
            normalized = {k: row.get(k) for k in fieldnames}
            writer.writerow(normalized)

    return len(rows)


def append_ledger_entries(entries: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """Append ledger rows directly into the XLSX for each account."""
    if not entries:
        return []

    appended_rows: List[Dict[str, Any]] = []
    for entry in entries:
        descriptor = resolve_account(entry.get("account"))
        wb = openpyxl.load_workbook(descriptor.path)
        ws = wb["Ledger"]
        _ensure_ledger_headers(ws)
        cols = _get_ledger_columns(ws)
        r = ws.max_row + 1

        ticker = str(entry.get("ticker") or "").upper()
        action = str(entry.get("action") or "Open").strip().upper()
        side = str(entry.get("side") or "Call").capitalize()
        contracts = int(entry.get("contracts"))
        strike = entry.get("strike")
        premium = entry.get("premium")
        underlying = entry.get("underlying")
        # Store expiry as a midnight datetime to keep Excel column typed as datetime
        expiry_ts = pd.to_datetime(entry.get("expiry"), errors="coerce").normalize()
        expiry = expiry_ts.to_pydatetime() if expiry_ts is not pd.NaT else None
        trade_dt = pd.to_datetime(entry.get("trade_datetime"), errors="coerce").to_pydatetime()

        account_label = descriptor.label
        if "christie" in descriptor.name.lower():
            account_label = "Christie"
        elif "travis" in descriptor.name.lower():
            account_label = "Travis"

        key = _build_key(ticker, strike, expiry, side, action)

        if "Account" in cols:
            ws.cell(row=r, column=cols["Account"], value=account_label)
        if "Date" in cols:
            ws.cell(row=r, column=cols["Date"], value=trade_dt)
        if "Action" in cols:
            ws.cell(row=r, column=cols["Action"], value=action)
        if "Symbol" in cols:
            ws.cell(row=r, column=cols["Symbol"], value=ticker)
        if "Contracts" in cols:
            ws.cell(row=r, column=cols["Contracts"], value=contracts)
        if "Strike" in cols:
            ws.cell(row=r, column=cols["Strike"], value=strike)
        if "Expiry" in cols:
            ws.cell(row=r, column=cols["Expiry"], value=expiry)
        if "Premium/Buyback" in cols:
            ws.cell(row=r, column=cols["Premium/Buyback"], value=premium)
        if "Underlying" in cols:
            if underlying in (None, ""):
                # Attempt to reuse latest underlying for this key if omitted
                try:
                    underlying = _latest_underlying(ws, cols, ticker, strike, expiry, side) or _cached_underlying(ticker, expiry, trade_dt)
                except Exception as exc:
                    raise ValueError(f"Unable to resolve underlying for {ticker} {strike} {expiry} {side}: {exc}")
            if underlying in (None, ""):
                raise ValueError(f"Missing underlying for {ticker} {strike} {expiry} {side}; aborting write")
            ws.cell(row=r, column=cols["Underlying"], value=underlying)
        if "Base Position Id" in cols:
            ws.cell(row=r, column=cols["Base Position Id"], value=entry.get("base_position_id"))
        if "Base Leg Id" in cols:
            ws.cell(row=r, column=cols["Base Leg Id"], value=entry.get("base_leg_id"))
        if "Key" in cols:
            ws.cell(row=r, column=cols["Key"], value=key)
        if "Side" in cols:
            ws.cell(row=r, column=cols["Side"], value=side)

        wb.save(descriptor.path)

        appended_rows.append(
            {
                "account": descriptor.name,
                "date": trade_dt.date().isoformat() if hasattr(trade_dt, "date") else None,
                "action": action,
                "side": side,
                "ticker": ticker,
                "contracts": contracts,
                "strike": strike,
        "expiry": expiry.isoformat() if expiry else None,
                "premium_buyback": premium,
                "underlying": underlying,
                "juice_per_contract": None,
                "signed_juice_dollars": None,
                "signed_juice_per_100": None,
                "key": key,
                "notes": None,
                "base_position_id": entry.get("base_position_id"),
                "base_leg_id": entry.get("base_leg_id"),
                "row_number": r,
            }
        )

    return appended_rows


def _get_ledger_columns(ws) -> Dict[str, int]:
    headers = {}
    for cell in ws[1]:
        if cell.value is None:
            continue
        headers[str(cell.value).strip().lower()] = cell.column

    aliases = {
        "account": "Account",
        "date": "Date",
        "action": "Action",
        "symbol": "Symbol",
        "contracts": "Contracts",
        "strike": "Strike",
        "expiry": "Expiry",
        "premium/buyback": "Premium/Buyback",
        "underlying": "Underlying",
        "key": "Key",
        "side": "Side",
        "base position id": "Base Position Id",
        "base leg id": "Base Leg Id",
    }
    return {target: headers[key] for key, target in aliases.items() if key in headers}


def _ensure_ledger_headers(ws) -> None:
    """Ensure legacy ledgers drop Condition and include Base Position/Leg Id."""
    headers = [cell.value for cell in ws[1]]
    changed = False
    if "Condition" in headers:
        idx = headers.index("Condition") + 1
        ws.delete_cols(idx)
        headers.pop(idx - 1)
        changed = True
    if "Base Position Id" not in headers:
        ws.cell(row=1, column=len(headers) + 1, value="Base Position Id")
        headers.append("Base Position Id")
        changed = True
    if "Base Leg Id" not in headers:
        ws.cell(row=1, column=len(headers) + 1, value="Base Leg Id")
        changed = True
    # Caller is responsible for saving the workbook; avoid using ws.parent.filename.


def _latest_underlying(ws, cols: Dict[str, int], ticker: str, strike: Any, expiry: Any, side: str) -> Any:
    """Find the most recent underlying price for matching ticker/strike/expiry/side."""
    try:
        ticker_col = cols.get("Symbol")
        strike_col = cols.get("Strike")
        expiry_col = cols.get("Expiry")
        underlying_col = cols.get("Underlying")
        side_col = cols.get("Side")
        if not (ticker_col and strike_col and expiry_col and underlying_col):
            return None
        target_expiry = _serialize_date(pd.to_datetime(expiry, errors="coerce"))
        for row in range(ws.max_row, 1, -1):
            tval = ws.cell(row=row, column=ticker_col).value
            sval = ws.cell(row=row, column=strike_col).value
            eval_val = ws.cell(row=row, column=expiry_col).value
            side_val = ws.cell(row=row, column=side_col).value if side_col else None
            if (
                str(tval or "").upper() == ticker
                and sval == strike
                and _serialize_date(pd.to_datetime(eval_val, errors="coerce")) == target_expiry
            ):
                if side_val and str(side_val).lower() not in str(side or "").lower():
                    continue
                return ws.cell(row=row, column=underlying_col).value
    except Exception:
        return None
    return None


def _refresh_alpha_cache(ticker: str, expiry: Any) -> None:
    """Fetch intraday (1min) prices from Alpha Vantage and refresh the cache file for the month."""
    api_key = os.environ.get("ALPHAVANTAGE_API_KEY")
    if not api_key:
        raise RuntimeError("ALPHAVANTAGE_API_KEY not set; cannot refresh alpha cache")
    try:
        ticker = ticker.upper()
        expiry_dt = pd.to_datetime(expiry, errors="coerce")
        if pd.isna(expiry_dt):
            return
        month_str = expiry_dt.strftime("%Y-%m")
        url = "https://www.alphavantage.co/query"
        params = {
            "function": "TIME_SERIES_INTRADAY",
            "symbol": ticker,
            "interval": "1min",
            "outputsize": "full",
            "apikey": api_key,
        }
        resp = requests.get(url, params=params, timeout=10)
        resp.raise_for_status()
        payload = resp.json()
        cache_file = ALPHA_CACHE_DIR / f"{ticker}_{month_str}.json"
        cache_file.parent.mkdir(parents=True, exist_ok=True)
        # Intraday payload key contains interval
        series = None
        for key in payload.keys():
            if key.lower().startswith("time series"):
                series = payload.get(key)
                break
        if series is None:
            series = {}
        # Filter to target month; intraday timestamps include time portion.
        # If the month is missing, cache the full series so lookups still work.
        filtered = {k: v for k, v in series.items() if str(k).startswith(month_str)}
        if not filtered:
            filtered = series
        cache_payload = {
            "__alpha_cache_format__": "raw",
            "raw": payload,
            "time_series": filtered,
        }
        cache_file.write_text(json.dumps(cache_payload, indent=2), encoding="utf-8")
        print(f"[alpha] cached response -> {cache_file}")
        if not series:
            raise RuntimeError("Alpha Vantage returned no data")
    except requests.RequestException as exc:
        raise RuntimeError(f"Alpha Vantage request failed: {exc}") from exc
    except Exception:
        raise


def _cached_underlying(ticker: str, expiry: Any, trade_dt: Any = None) -> Any:
    """Attempt to pull the latest underlying from alpha_cache, refreshing if stale."""
    def _unwrap_alpha_cache_payload(payload: Any) -> Dict[str, Any]:
        if isinstance(payload, dict):
            if payload.get("__alpha_cache_format__") == "raw":
                series = payload.get("time_series")
                if isinstance(series, dict):
                    return series
            for key, value in payload.items():
                if isinstance(key, str) and key.lower().startswith("time series") and isinstance(value, dict):
                    return value
        return payload if isinstance(payload, dict) else {}

    ticker = ticker.upper()
    expiry_dt = pd.to_datetime(expiry, errors="coerce")
    trade_ts = pd.to_datetime(trade_dt, errors="coerce") if trade_dt is not None else None
    # Pick cache month from trade date if available; fallback to expiry.
    target_dt = trade_ts if trade_ts is not None and not pd.isna(trade_ts) else expiry_dt
    if pd.isna(target_dt):
        return None
    month_str = target_dt.strftime("%Y-%m")
    fname = f"{ticker}_{month_str}.json"
    path = ALPHA_CACHE_DIR / fname
    refresh_needed = not path.exists()
    data = {}
    if path.exists():
        with path.open("r", encoding="utf-8") as f:
            try:
                data = json.load(f)
            except Exception:
                refresh_needed = True
        data = _unwrap_alpha_cache_payload(data)
    if not isinstance(data, dict):
        refresh_needed = True
    if isinstance(data, dict) and not data:
        refresh_needed = True

    trade_ts = trade_ts  # keep name stable below
    # Alpha Vantage intraday timestamps are US/Eastern; user inputs are treated as local.
    # Empirically the feed is ~1 hour ahead of the entered trade time, so bias lookup by +1h.
    trade_ts_lookup = trade_ts + pd.Timedelta(hours=1) if trade_ts is not None and not pd.isna(trade_ts) else None
    latest_ts = max(data.keys()) if data else None
    if latest_ts and trade_ts_lookup is not None:
        if pd.to_datetime(latest_ts) < trade_ts_lookup:
            refresh_needed = True

    refresh_error: Exception | None = None
    if refresh_needed:
        try:
            _refresh_alpha_cache(ticker, expiry_dt)
        except Exception as exc:
            refresh_error = exc
        if path.exists():
            with path.open("r", encoding="utf-8") as f:
                data = json.load(f)
            data = _unwrap_alpha_cache_payload(data)
            if not isinstance(data, dict):
                refresh_error = RuntimeError(f"Alpha cache malformed for {ticker} {month_str}")
        else:
            refresh_error = refresh_error or RuntimeError(f"Alpha cache refresh failed for {ticker} {month_str}")
        # If refresh succeeded but the cache file is still missing, persist what we have.
        if refresh_error is None and data:
            path.parent.mkdir(parents=True, exist_ok=True)
            path.write_text(json.dumps(data, indent=2), encoding="utf-8")
        # If refresh failed, fall back to any cached month for this ticker.
        if refresh_error is not None:
            fallback = None
            for candidate in sorted(ALPHA_CACHE_DIR.glob(f"{ticker}_*.json"), reverse=True):
                try:
                    with candidate.open("r", encoding="utf-8") as f:
                        loaded = json.load(f)
                    loaded = _unwrap_alpha_cache_payload(loaded)
                    if isinstance(loaded, dict) and loaded:
                        fallback = loaded
                        break
                except Exception:
                    continue
            if fallback is None:
                raise RuntimeError(str(refresh_error))
            data = fallback

    timestamps = sorted(data.keys())
    if not timestamps:
        raise RuntimeError(f"No alpha cache data for {ticker} {month_str}")

    chosen_ts = None
    if trade_ts_lookup is not None:
        for ts in reversed(timestamps):
            ts_dt = pd.to_datetime(ts)
            if ts_dt <= trade_ts_lookup:
                chosen_ts = ts
                break
    if not chosen_ts:
        chosen_ts = timestamps[-1]

    # If we still could not find a price at or before trade_ts, attempt one more refresh for safety.
    if trade_ts_lookup is not None:
        chosen_dt = pd.to_datetime(chosen_ts)
        if chosen_dt < trade_ts_lookup:
            _refresh_alpha_cache(ticker, expiry_dt)
            with path.open("r", encoding="utf-8") as f:
                data = json.load(f)
            data = _unwrap_alpha_cache_payload(data)
            timestamps = sorted(data.keys())
            for ts in reversed(timestamps):
                ts_dt = pd.to_datetime(ts)
                if ts_dt <= trade_ts_lookup:
                    chosen_ts = ts
                    break

    entry = data.get(chosen_ts, {})
    # Prefer mid of open/close when available
    def _to_float(val):
        try:
            return float(val)
        except Exception:
            return None

    open_val = _to_float(entry.get("1. open") or entry.get("open"))
    close_val = _to_float(entry.get("4. close") or entry.get("close") or entry.get("price"))
    high_val = _to_float(entry.get("2. high") or entry.get("high"))
    low_val = _to_float(entry.get("3. low") or entry.get("low"))

    ohlc = [v for v in (open_val, close_val, high_val, low_val) if v is not None]
    if len(ohlc) >= 2:
        return round(sum(ohlc) / len(ohlc), 4)
    if close_val is not None:
        return close_val
    if open_val is not None:
        return open_val

    raise RuntimeError(f"Alpha cache entry missing open/close price for {ticker} {chosen_ts}")


def _build_key(ticker: str, strike: Any, expiry: Any, side: str, action: str) -> str:
    expiry_str = ""
    if hasattr(expiry, "isoformat"):
        try:
            expiry_str = expiry.date().isoformat()
        except Exception:
            expiry_str = expiry.isoformat()
    elif expiry is not None:
        expiry_str = str(expiry)
    return f"{ticker}|{strike}|{expiry_str}|{side.upper()}|{action.upper()}".upper()


def update_ledger_row(account: str, row_number: int, data: Dict[str, Any]) -> Dict[str, Any]:
    """Update an existing ledger row in the Excel workbook and return the updated record."""
    descriptor = _resolve_account(account)
    wb = openpyxl.load_workbook(descriptor.path)
    ws = wb["Ledger"]
    _ensure_ledger_headers(ws)
    cols = _get_ledger_columns(ws)

    r = int(row_number)
    if r < 2 or r > ws.max_row:
        raise ValueError(f"Row {r} is out of range for ledger {descriptor.path.name}")

    account_label = descriptor.label
    if "christie" in account_label.lower():
        account_label = "Christie"
    elif "travis" in account_label.lower():
        account_label = "Travis"

    # Write core fields
    if "Account" in cols:
        ws.cell(row=r, column=cols["Account"], value=account_label)
    if "Date" in cols:
        ws.cell(row=r, column=cols["Date"], value=data.get("trade_datetime"))
    action_val = str(data.get("action") or "Open").strip().upper()
    if "Action" in cols:
        ws.cell(row=r, column=cols["Action"], value=action_val)
    if "Symbol" in cols:
        ws.cell(row=r, column=cols["Symbol"], value=(data.get("ticker") or "").upper())
    if "Contracts" in cols:
        ws.cell(row=r, column=cols["Contracts"], value=data.get("contracts"))
    if "Strike" in cols:
        ws.cell(row=r, column=cols["Strike"], value=data.get("strike"))
    if "Expiry" in cols:
        expiry_ts = pd.to_datetime(data.get("expiry"), errors="coerce").normalize()
        expiry_val = expiry_ts.to_pydatetime() if expiry_ts is not pd.NaT else None
        ws.cell(row=r, column=cols["Expiry"], value=expiry_val)
    if "Premium/Buyback" in cols:
        ws.cell(row=r, column=cols["Premium/Buyback"], value=data.get("premium"))
    if "Underlying" in cols:
        underlying = data.get("underlying")
        if underlying in (None, ""):
            try:
                underlying = _latest_underlying(
                    ws,
                    cols,
                    (data.get("ticker") or "").upper(),
                    data.get("strike"),
                    pd.to_datetime(data.get("expiry"), errors="coerce"),
                    data.get("side"),
                ) or _cached_underlying((data.get("ticker") or "").upper(), data.get("expiry"), data.get("trade_datetime"))
            except Exception as exc:
                raise ValueError(f"Unable to resolve underlying for {data.get('ticker')} {data.get('strike')} {data.get('expiry')} {data.get('side')}: {exc}")
        if underlying in (None, ""):
            raise ValueError(f"Missing underlying for {data.get('ticker')} {data.get('strike')} {data.get('expiry')} {data.get('side')}; aborting update")
        ws.cell(row=r, column=cols["Underlying"], value=underlying)
    if "Base Position Id" in cols:
        ws.cell(row=r, column=cols["Base Position Id"], value=data.get("base_position_id"))
    if "Base Leg Id" in cols:
        ws.cell(row=r, column=cols["Base Leg Id"], value=data.get("base_leg_id"))
    if "Side" in cols:
        ws.cell(row=r, column=cols["Side"], value=data.get("side") or "Call")
    if "Key" in cols:
        key = _build_key(
            (data.get("ticker") or "").upper(),
            data.get("strike"),
            data.get("expiry"),
            data.get("side") or "Call",
            data.get("action") or "Open",
        )
        ws.cell(row=r, column=cols["Key"], value=key)
    wb.save(descriptor.path)

    # Return the updated row as a dict matching LedgerRow schema
    updated = {
        "account": descriptor.name,
        "date": data.get("trade_datetime"),
        "action": data.get("action"),
        "side": data.get("side"),
        "ticker": (data.get("ticker") or "").upper(),
        "contracts": data.get("contracts"),
        "strike": data.get("strike"),
        "expiry": _serialize_date(pd.to_datetime(data.get("expiry"), errors="coerce")),
        "premium_buyback": data.get("premium"),
        "underlying": data.get("underlying"),
        "juice_per_contract": None,
        "signed_juice_dollars": None,
        "signed_juice_per_100": None,
        "key": None,
        "notes": None,
        "row_number": r,
        "base_position_id": data.get("base_position_id"),
        "base_leg_id": data.get("base_leg_id"),
    }
    return updated


def get_ledger_rows(account: str | None = None) -> List[Dict[str, Any]]:
    """Load raw ledger rows from the XLSX for display in the UI."""
    descriptor = _resolve_account(account)
    # Ensure headers up to date
    try:
        wb = openpyxl.load_workbook(descriptor.path)
        ws = wb["Ledger"]
        _ensure_ledger_headers(ws)
        wb.save(descriptor.path)
    except Exception:
        pass

    df = pd.read_excel(descriptor.path, engine="openpyxl")
    if df.empty:
        return []

    df.columns = _normalize_columns(df.columns.tolist())
    df = df.replace({np.nan: None})

    # Drop completely empty rows (no ticker)
    if "ticker" in df.columns:
        df = df[df["ticker"].astype(str).str.strip() != ""]

    # Excel row numbers (data starts at row 2 because row 1 is headers)
    df["row_number"] = df.index + 2

    if "date" in df.columns:
        df["date"] = pd.to_datetime(df["date"], errors="coerce")

    rename_map = {
        "symbol": "ticker",
        "premium/buyback": "premium_buyback",
    }
    for src, dest in rename_map.items():
        if src in df.columns:
            df[dest] = df[src]

    numeric_cols = [
        "contracts",
        "strike",
        "premium_buyback",
        "underlying",
    ]
    for col in numeric_cols:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce")

    # Heuristic fix: if side is blank but a side-like value leaked into signed_juice columns, rescue it.
    if "side" in df.columns and "signed_juice_dollars" in df.columns:
        mask = df["side"].isna() & df["signed_juice_dollars"].astype(str).str.lower().isin(["call", "put"])
        if mask.any():
            df.loc[mask, "side"] = df.loc[mask, "signed_juice_dollars"]
            df.loc[mask, "signed_juice_dollars"] = None
            if "signed_juice_per_100" in df.columns:
                df.loc[mask, "signed_juice_per_100"] = None

    df["account"] = descriptor.name
    if "contracts" in df.columns:
        df["contracts"] = pd.to_numeric(df["contracts"], errors="coerce").astype("Int64")
    if "expiry" in df.columns:
        df["expiry"] = pd.to_datetime(df["expiry"], errors="coerce")

    columns = [
        "account",
        "date",
        "action",
        "side",
        "ticker",
        "contracts",
        "strike",
        "expiry",
        "premium_buyback",
        "underlying",
        "base_position_id",
        "base_leg_id",
        "notes",
        "condition",
        "key",
        "row_number",
    ]

    available = [col for col in columns if col in df.columns]
    records = df[available].to_dict("records")
    # Ensure all keys exist for consumers
    normalized_records: List[Dict[str, Any]] = []
    for record in records:
        normalized: Dict[str, Any] = {col: record.get(col) for col in columns}
        # Prefer explicit condition column; fall back to notes if absent
        if not normalized.get("condition"):
            normalized["condition"] = normalized.get("notes")
        # Clean non-finite numeric values
        for col in [
            "contracts",
            "strike",
            "premium_buyback",
            "underlying",
        ]:
            val = normalized.get(col)
            try:
                if val is None:
                    continue
                if not pd.notna(val) or not np.isfinite(float(val)):
                    normalized[col] = None
            except Exception:
                normalized[col] = None
        # Recover side if it leaked into juice columns
        if not normalized.get("side"):
            for candidate in ["signed_juice_dollars", "signed_juice_per_100", "juice_per_contract"]:
                val = normalized.get(candidate)
                if isinstance(val, str) and val.strip().lower() in {"call", "put"}:
                    normalized["side"] = val.strip().title()
                    normalized[candidate] = None
                    break
        # Compute key if missing
        if not normalized.get("key"):
            normalized["key"] = _composite_key(
                normalized.get("ticker"),
                normalized.get("strike"),
                normalized.get("expiry"),
                normalized.get("side"),
                normalized.get("action"),
            )
        exp_val = normalized.get("expiry")
        if hasattr(exp_val, "isoformat"):
            try:
                normalized["expiry"] = exp_val.date().isoformat() if hasattr(exp_val, "date") else exp_val.isoformat()
            except Exception:
                normalized["expiry"] = exp_val.isoformat()
        def r2(val):
            return round(val, 2) if isinstance(val, (int, float)) else val
        contracts_val = normalized.get("contracts")
        if contracts_val is not None:
            try:
                normalized["contracts"] = int(contracts_val)
            except (TypeError, ValueError):
                normalized["contracts"] = None
        date_val = normalized.get("date")
        if hasattr(date_val, "isoformat"):
            try:
                normalized["date"] = date_val.isoformat()
            except Exception:
                normalized["date"] = None
        elif pd.isna(date_val):
            normalized["date"] = None
        expiry_val = normalized.get("expiry")
        if hasattr(expiry_val, "isoformat"):
            try:
                normalized["expiry"] = expiry_val.isoformat()
            except Exception:
                normalized["expiry"] = None
        elif pd.isna(expiry_val):
            normalized["expiry"] = None
        # Keep calculated fields as-is from the ledger (UI will compute display values).
        normalized["juice_per_contract"] = r2(normalized.get("juice_per_contract"))
        normalized["signed_juice_dollars"] = r2(normalized.get("signed_juice_dollars"))
        normalized["signed_juice_per_100"] = r2(normalized.get("signed_juice_per_100"))
        if normalized.get("row_number") is not None:
            try:
                normalized["row_number"] = int(normalized["row_number"])
            except (TypeError, ValueError):
                normalized["row_number"] = None
        else:
            normalized["row_number"] = None
        normalized_records.append(normalized)
    return normalized_records


def get_trades_for_current_week(account: str | None = None) -> pd.DataFrame:
    trades = get_all_trades(account)
    if trades.empty:
        return trades

    today = pd.Timestamp.now().normalize()
    start_of_week = today - pd.to_timedelta(today.weekday(), unit="D")
    end_of_week = start_of_week + pd.Timedelta(days=7)

    mask = (trades["date"] >= start_of_week) & (trades["date"] < end_of_week)
    return trades.loc[mask].copy()


def get_weekly_summary(account: str | None = None) -> pd.DataFrame:
    trades = get_all_trades(account)
    if trades.empty:
        return trades

    trades = trades.copy()
    trades["date"] = pd.to_datetime(trades["date"], errors="coerce")

    working = trades.copy()
    working["week_start"] = (
        working["date"] - pd.to_timedelta(working["date"].dt.weekday, unit="D")
    )

    summary = (
        working.groupby(["week_start", "strategy"], as_index=False)
        .agg(
            total_juice=("juice", "sum"),
            total_basis=("basis", "sum"),
            trade_count=("ticker", "count"),
        )
        .sort_values(["week_start", "strategy"])
    )

    return summary


def get_dashboard_metrics(account: str | None = None) -> Dict[str, Any]:
    trades = get_all_trades(account)
    current_week = get_trades_for_current_week(account)

    cumulative = float(trades["juice"].sum()) if not trades.empty else 0.0
    total_trades = len(trades)

    return {
        "total_trades": total_trades,
        "cumulative_juice": cumulative,
        "first_trade_date": _serialize_date(trades["date"].min()),
        "last_trade_date": _serialize_date(trades["date"].max()),
        "current_week_trade_count": len(current_week),
    }
